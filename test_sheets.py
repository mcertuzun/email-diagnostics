# -*- coding: utf-8 -*-
"""
Tests for the Excel presentation: sorting, colouring, autofilter and the
Summary and Companies sheets.

Companies is a different grain from Results, one row per company rather than
per contact, so it must not duplicate rows. Sorting by action must not lose
traceability, which is what source_row is for.

No network access.
"""
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
os.chdir(HERE)

import openpyxl
import email_diagnostics as ED


class Resp(object):
    def __init__(self, payload):
        self._payload, self.status_code, self.headers = payload, 200, {}

    def json(self):
        return self._payload


REGISTRY = {
    "00000001": [
        {"name": "SMITH, John", "name_elements": {"forename": "John", "surname": "Smith"},
         "officer_role": "director"},
        {"name": "PATEL, Priya", "name_elements": {"forename": "Priya", "surname": "Patel"},
         "officer_role": "secretary"},
    ],
    "00000002": [
        {"name": "OLD, Gone", "name_elements": {"forename": "Gone", "surname": "Old"},
         "officer_role": "director", "resigned_on": "2021-05-01"},
        {"name": "NEW, Nina", "name_elements": {"forename": "Nina", "surname": "New"},
         "officer_role": "director"},
    ],
}


def fake_get(url, params=None, timeout=None):
    if url.endswith("officers"):
        number = url.rstrip("/").split("/")[-2]
        items = REGISTRY.get(number, [])
        return Resp({"items": items, "total_results": len(items),
                     "items_per_page": 100, "start_index": 0})
    number = url.split("/")[-1]
    return Resp({"company_name": "ACME TRADING LIMITED" if number == "00000001"
                 else "BETA HOLDINGS LIMITED",
                 "company_status": "active"})


# Deliberately in an order that sorting must change
ROWS = [
    ("John", "Smith", "john.smith@acme.co.uk", "Acme Trading Ltd", "00000001"),
    ("Gone", "Old", "gone.old@beta.co.uk", "Beta Holdings Ltd", "00000002"),
    ("John", "Smith", "jhon.smith@acme.co.uk", "Acme Trading Ltd", "00000001"),
    ("Priya", "Patel", "priya.patel@acme.co.uk", "Acme Trading Ltd", "00000001"),
    ("Nobody", "Here", "nobody@beta.co.uk", "Beta Holdings Ltd", "00000002"),
]

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["first_name", "last_name", "email", "company", "regnum", "status"])
for first, last, email, company, regnum in ROWS:
    sheet.append([first, last, email, company, regnum, "Bounced"])
workbook.save("sh_in.xlsx")

original = ED.CompaniesHouseClient


class Patched(original):
    def __init__(self, *a, **k):
        original.__init__(self, *a, **k)
        self._session.get = fake_get


ED.CompaniesHouseClient = Patched
os.environ["CH_API_KEY"] = "fake"
ED.cli(["triage", "-i", "sh_in.xlsx", "-o", "sh_out.xlsx"])
ED.CompaniesHouseClient = original

book = openpyxl.load_workbook("sh_out.xlsx")
failures = []


def check(label, got, want):
    ok = got == want
    if not ok:
        failures.append("%s -> got %r want %r" % (label, got, want))
    print("%-6s %-56s %s" % ("OK" if ok else "FAIL", label, got))


print("=" * 92)
print(" SHEETS")
print("=" * 92)
print("     %s" % book.sheetnames)
check("Results opens first", book.sheetnames[0], "Results")
check("Summary and Companies present",
      "Summary" in book.sheetnames and "Companies" in book.sheetnames, True)

worklists = [n for n in book.sheetnames if n not in ("Results", "Summary", "Companies")]
check("row counts in the tab names", all("(" in n for n in worklists), True)
check("empty queues get no sheet",
      any(n.startswith("Fix data") for n in book.sheetnames), False)

results = book["Results"]
rows = list(results.iter_rows(values_only=True))
header, data = list(rows[0]), [list(r) for r in rows[1:]]
action_index = header.index("action")
source_index = header.index("source_row")

print()
print("=" * 92)
print(" RESULTS SHEET")
print("=" * 92)
check("one row per input row", len(data), len(ROWS))
check("frozen header", results.freeze_panes, "A2")
check("autofilter set", results.auto_filter.ref is not None, True)

actions = [r[action_index] for r in data]
print("     order after sorting: %s" % actions)
rank = dict((name, i) for i, name in enumerate(ED.ACTION_ORDER))
check("sorted by ACTION_ORDER", actions == sorted(actions, key=lambda a: rank[a]), True)
check("all-correct rows sit last", actions[-1], ED.A.ALL_CORRECT)

check("source_row present for every row",
      all(isinstance(r[source_index], int) for r in data), True)
check("source_row values are the input positions",
      sorted(r[source_index] for r in data), list(range(2, 2 + len(ROWS))))

coloured = 0
for row in range(2, len(data) + 2):
    cell = results.cell(row=row, column=action_index + 1)
    if cell.fill is not None and cell.fill.fgColor.rgb not in (None, "00000000"):
        coloured += 1
check("action cell coloured on every row", coloured, len(data))
check("header is bold", results.cell(row=1, column=1).font.bold, True)

print()
print("=" * 92)
print(" SUMMARY SHEET")
print("=" * 92)
summary_text = "\n".join(
    " ".join(str(v) for v in row if v is not None)
    for row in book["Summary"].iter_rows(values_only=True))
for token in ("ACTION", "RESULT", "REASON", "rows analysed", "HTTP requests", "mode"):
    check("mentions %r" % token, token in summary_text, True)

print()
print("=" * 92)
print(" COMPANIES SHEET")
print("=" * 92)
companies = book["Companies"]
crows = list(companies.iter_rows(values_only=True))
cheader, cdata = list(crows[0]), [list(r) for r in crows[1:]]
print("     %s" % ", ".join(str(h) for h in cheader))
for row in cdata:
    print("     %s" % row)

check("one row per company, not per contact", len(cdata), 2)
check("distinct regnums", sorted(r[0] for r in cdata), ["00000001", "00000002"])
contacts_index = cheader.index("bounced_contacts")
check("contact counts sum to the row count",
      sum(r[contacts_index] for r in cdata), len(ROWS))
gone_index = cheader.index("resigned_or_missing")
check("company with departures sorted first", cdata[0][0], "00000002")
check("  and it counts both of them", cdata[0][gone_index], 2)
sugg_index = cheader.index("active_officer_suggestions")
check("suggestions carried through", "Nina New" in (cdata[0][sugg_index] or ""), True)

print()
print("=" * 92)
print(" WORK QUEUE SHEETS")
print("=" * 92)
for name in worklists:
    sheet = book[name]
    rows = list(sheet.iter_rows(values_only=True))
    print("     %-28s %s rows   %s" % (name, len(rows) - 1,
                                       ", ".join(str(c) for c in rows[0])))

fix_name = [n for n in worklists if n.startswith("Fix address")][0]
fix_rows = list(book[fix_name].iter_rows(values_only=True))
check("Fix address holds only its own rows", len(fix_rows) - 1, 1)
check("narrow view, not a copy of Results",
      len(fix_rows[0]) < len(header), True)
check("source_row is the first column", fix_rows[0][0], "source_row")

find_name = [n for n in worklists if n.startswith("Find new contact")][0]
find_rows = list(book[find_name].iter_rows(values_only=True))
check("Find new contact row count", len(find_rows) - 1, 2)
find_header = list(find_rows[0])
check("carries the suggestions column",
      "active_officer_suggestions" in find_header, True)
check("does not carry email pattern noise", "result_reason" in find_header, False)

check("no Investigate sheet when nothing needs looking at",
      any(n.startswith("Investigate") for n in book.sheetnames), False)

check("Fix address carries the input company", "company" in list(fix_rows[0]), True)
check("Fix address carries the registered name",
      "companyhouse_names" in list(fix_rows[0]), True)
company_index = list(fix_rows[0]).index("company")
check("  and it is populated", fix_rows[1][company_index], "Acme Trading Ltd")

all_correct_rows = [r for r in data if r[action_index] == ED.A.ALL_CORRECT]
no_action_name = [n for n in worklists if n.startswith("No action")][0]
no_action_rows = list(book[no_action_name].iter_rows(values_only=True))
check("No action sheet exists", no_action_name.startswith("No action"), True)
check("holds every all-correct row", len(no_action_rows) - 1, len(all_correct_rows))
check("all-correct is NOT in the Investigate queue",
      any(n.startswith("Investigate") for n in worklists), False)

in_queues = sum(len(list(book[name].iter_rows(values_only=True))) - 1
                for name in worklists)
check("every row lands in exactly one queue", in_queues, len(data))

os.remove("sh_in.xlsx")
os.remove("sh_out.xlsx")

# --plain must switch all of it off
ED.CompaniesHouseClient = Patched
workbook.save("sh_in.xlsx")
ED.cli(["triage", "-i", "sh_in.xlsx", "-o", "sh_plain.xlsx", "--plain"])
ED.CompaniesHouseClient = original
plain = openpyxl.load_workbook("sh_plain.xlsx")
print()
print("=" * 92)
print(" --plain")
print("=" * 92)
check("single sheet", plain.sheetnames, ["Results"])
plain_rows = list(plain["Results"].iter_rows(values_only=True))
plain_source = [r[list(plain_rows[0]).index("source_row")] for r in plain_rows[1:]]
check("input order kept", plain_source, list(range(2, 2 + len(ROWS))))
os.remove("sh_in.xlsx")
os.remove("sh_plain.xlsx")

# CSV output cannot carry any of it, and must say so rather than silently
# producing a flat file when sheets were expected.
print()
print("=" * 92)
print(" CSV OUTPUT")
print("=" * 92)
workbook.save("sh_in.xlsx")
ED.CompaniesHouseClient = Patched
messages = []


class Capture(object):
    def __init__(self, sink):
        self.sink = sink

    def handle(self, record):
        self.sink.append(record.getMessage())

    def level(self):
        return 0


import logging


class _Handler(logging.Handler):
    def emit(self, record):
        messages.append(record.getMessage())


handler = _Handler()
ED.log.addHandler(handler)
ED.cli(["triage", "-i", "sh_in.xlsx", "-o", "sh_out.csv"])
ED.log.removeHandler(handler)
ED.CompaniesHouseClient = original

joined = "\n".join(messages)
check("warns that CSV holds one table", "ONE flat table" in joined, True)
check("names the .xlsx alternative", ".xlsx" in joined, True)
with io.open("sh_out.csv", encoding="utf-8-sig") as handle:
    csv_header = handle.readline().strip().split(",")
check("action column still present in CSV", "action" in csv_header, True)
check("source_row still present in CSV", "source_row" in csv_header, True)
os.remove("sh_in.xlsx")
os.remove("sh_out.csv")

print()
print("=" * 92)
if failures:
    print("%d FAILURE(S):" % len(failures))
    for item in failures:
        print("   -", item)
    sys.exit(1)
print("SHEET TESTS PASSED")
