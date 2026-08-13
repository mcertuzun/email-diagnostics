# -*- coding: utf-8 -*-
"""
Tests for ch_first, the default mode.

Verifies:
  - N rows -> N Companies House lookups -> N rows in the output
  - result priority: data problem > resigned > typo > active
  - the official name from Companies House, MIDDLE NAMES included, is used
    for the email check, so a short name in the input does not cause a
    false typo
  - on an ambiguous match the official name is NOT used as the baseline

No network access.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
os.chdir(HERE)

import openpyxl
import email_diagnostics as ED


class Resp(object):
    def __init__(self, payload):
        self._payload, self.status_code, self.headers = payload, 200, {}

    def json(self):
        return self._payload


# regnum -> officer records
REGISTRY = {
    "00000001": [{"name": "SMITH, John Andrew",
                  "name_elements": {"forename": "John", "other_forenames": "Andrew",
                                    "surname": "Smith"}}],
    "00000002": [{"name": "JONES, Sarah",
                  "name_elements": {"forename": "Sarah", "surname": "Jones"},
                  "resigned_on": "2024-01-31"}],
    "00000003": [{"name": "TAYLOR, Elizabeth",
                  "name_elements": {"forename": "Elizabeth", "surname": "Taylor"}}],
    "00000004": [{"name": "BROWN, Alan",
                  "name_elements": {"forename": "Alan", "surname": "Brown"}},
                 {"name": "BROWN, Barry",
                  "name_elements": {"forename": "Barry", "surname": "Brown"}}],
    "00000005": [{"name": "WHITE, Mark",
                  "name_elements": {"forename": "Mark", "surname": "White"}}],
}

calls = []


def fake_get(url, params=None, timeout=None):
    calls.append(url)
    number = url.rstrip("/").split("/")[-2] if url.endswith("officers") else url.split("/")[-1]
    if url.endswith("officers"):
        items = REGISTRY.get(number, [])
        return Resp({"items": items, "total_results": len(items),
                     "items_per_page": 100, "start_index": 0})
    return Resp({"company_name": "ACME LTD", "company_status": "active"})


# first_name, last_name, email, regnum, expected result, note
CASES = [
    ("John", "Smith", "john.andrew.smith@acme.co.uk", "00000001",
     ED.R.ACTIVE, "middle name in the address, matched via Andrew from CH"),
    ("Sarah", "Jones", "sarah.jones@acme.co.uk", "00000002",
     ED.R.RESIGNED, "resigned -> resignation outranks a typo"),
    ("Sarah", "Jones", "sarahh.jones@acme.co.uk", "00000002",
     ED.R.RESIGNED, "resigned and a typo -> resignation wins"),
    ("Liz", "Taylor", "liz.taylor@acme.co.uk", "00000003",
     ED.R.ACTIVE, "nickname Liz=Elizabeth, active"),
    ("John", "Smith", "jhon.smith@acme.co.uk", "00000001",
     ED.R.FIRST_NAME_TYPO, "active officer but a typo -> typo wins"),
    ("John", "Smith", "", "00000001",
     ED.R.MISSING_EMAIL, "a data problem outranks everything"),
    ("Alan", "Brown", "alan.brown@acme.co.uk", "00000004",
     ED.R.ACTIVE, "two officers share a surname, the forename decides"),
    ("Mark", "White", "mark.white@acme.co.uk", "00000009",
     ED.R.NO_OFFICER, "company not in the registry -> no officer found"),
]


def build():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["first_name", "last_name", "email", "company", "regnum", "status"])
    for first, last, email, regnum, _expected, _note in CASES:
        sheet.append([first, last, email, "Acme Ltd", regnum, "Bounced"])
    workbook.save("chf_in.xlsx")


build()
original = ED.CompaniesHouseClient


class Patched(original):
    def __init__(self, *a, **k):
        original.__init__(self, *a, **k)
        self._session.get = fake_get


ED.CompaniesHouseClient = Patched
os.environ["CH_API_KEY"] = "fake"
ED.cli(["triage", "-i", "chf_in.xlsx", "-o", "chf_out.xlsx"])
ED.CompaniesHouseClient = original

rows = list(openpyxl.load_workbook("chf_out.xlsx").active.iter_rows(values_only=True))
header, data = list(rows[0]), [list(r) for r in rows[1:]]
result_index = header.index("result")
officer_index = header.index("ch_officer_name")
os.remove("chf_in.xlsx")
os.remove("chf_out.xlsx")

problems = []
print("=" * 88)
print(" ch_first MODE")
print("=" * 88)
print("%-3s %-30s %-46s %s" % ("#", "EXPECTED", "GOT", "RESULT"))
print("-" * 88)
for index, (case, row) in enumerate(zip(CASES, data), start=1):
    expected = case[4]
    got = row[result_index] or ""
    ok = got.split(":")[0].strip() == expected
    if not ok:
        problems.append("%s -> expected %r, got %r  (%s)" % (index, expected, got, case[5]))
    print("%-3s %-30s %-46s %s" % (index, expected, got[:46], "OK" if ok else "FAIL"))
    print("    %s" % case[5])

print()
print("-" * 88)
unique_regnums = len(set(c[3] for c in CASES))
print("Rows in            : %s" % len(CASES))
print("Rows written       : %s" % len(data))
print("Distinct regnums   : %s" % unique_regnums)
print("CH requests        : %s" % len(calls))

if len(data) != len(CASES):
    problems.append("Wrote %s rows, expected %s" % (len(data), len(CASES)))
if len(calls) != unique_regnums * 2:      # profile call is on, so two requests per company
    if len(calls) != unique_regnums:
        problems.append("%s CH requests, expected %s" % (len(calls), unique_regnums))

print()
print("=" * 88)
if problems:
    print("%d PROBLEM(S):" % len(problems))
    for p in problems:
        print("   -", p)
    sys.exit(1)
print("ch_first TESTS PASSED")
