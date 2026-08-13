# -*- coding: utf-8 -*-
"""
Tests for the acronym domain rule and the two added output columns.

Covers:
  - a domain made of the company acronym plus a company word is accepted
  - the acronym rule only applies to names of two or more words
  - companyhouse_names carries the official name and any former names
  - active_officer_suggestions lists who is in post when nobody matched

No network access.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
os.chdir(HERE)

import openpyxl
import email_diagnostics as ED

failures = []


def check(label, got, want):
    ok = got == want
    if not ok:
        failures.append("%s -> got %r want %r" % (label, got, want))
    print("%-6s %-58s %s" % ("OK" if ok else "FAIL", label, got))


print("=" * 84)
print(" 1) ACRONYM DOMAIN RULE   company: 'Ali Veli Zeynep Trading Ltd'")
print("=" * 84)

tokens = ED.clean_company_name("Ali Veli Zeynep Trading Ltd")
candidates = ED.generate_company_domain_candidates(tokens)
print("     cleaned tokens: %s" % tokens)

for domain, want in [
    ("avz.co.uk", "ok"),
    ("avzltd.co.uk", "ok"),
    ("avzgroup.com", "ok"),
    ("avzuk.com", "ok"),
    ("avzholdings.com", "ok"),
    ("avzsolutions.co.uk", "ok"),
    ("alivelizeynep.com", "ok"),
    ("aliveli.com", "ok"),
    ("xyzgroup.com", "unmatched"),      # unrelated acronym stays unmatched
]:
    verdict, _reason = ED.check_domain(ED.parse_email("a@" + domain), candidates)
    check("%-24s" % domain, verdict, want)

print()
print("-" * 84)
print(" single-word company must NOT get the acronym rule")
print("-" * 84)
single = ED.generate_company_domain_candidates(ED.clean_company_name("Acme Ltd"))
print("     cleaned tokens: %s" % ED.clean_company_name("Acme Ltd"))
check("acme.co.uk still matches", ED.check_domain(ED.parse_email("a@acme.co.uk"), single)[0], "ok")
check("'a'+suffix not a candidate", any(c in ("aco", "altd", "agroup") for c in single), False)

print()
print("=" * 84)
print(" 2) NEW OUTPUT COLUMNS")
print("=" * 84)


class Resp(object):
    def __init__(self, payload, code=200):
        self._payload, self.status_code, self.headers = payload, code, {}

    def json(self):
        return self._payload


OFFICERS = {
    "00000001": [
        {"name": "SMITH, John", "name_elements": {"forename": "John", "surname": "Smith"},
         "officer_role": "director"},
        {"name": "PATEL, Priya", "name_elements": {"forename": "Priya", "surname": "Patel"},
         "officer_role": "secretary"},
        {"name": "REID, Alan", "name_elements": {"forename": "Alan", "surname": "Reid"},
         "officer_role": "director"},
        {"name": "OLD, Gone", "name_elements": {"forename": "Gone", "surname": "Old"},
         "officer_role": "director", "resigned_on": "2020-01-01"},
    ],
}
PROFILES = {
    "00000001": {"company_name": "AVZ TRADING LIMITED", "company_status": "active",
                 # A former name is present but must NOT reach the output column.
                 "previous_company_names": [{"name": "ALI VELI ZEYNEP LIMITED",
                                             "effective_from": "2010-01-01",
                                             "ceased_on": "2018-06-30"}]},
}


def fake_get(url, params=None, timeout=None):
    if url.endswith("officers"):
        number = url.rstrip("/").split("/")[-2]
        items = OFFICERS.get(number, [])
        return Resp({"items": items, "total_results": len(items),
                     "items_per_page": 100, "start_index": 0})
    number = url.split("/")[-1]
    return Resp(PROFILES.get(number, {"company_name": "UNKNOWN LTD",
                                      "company_status": "active"}))


workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["first_name", "last_name", "email", "company", "regnum", "status"])
# row 1: matches an officer
sheet.append(["John", "Smith", "john.smith@avzgroup.com", "Ali Veli Zeynep Ltd",
              "00000001", "Bounced"])
# row 2: nobody by this name -> should get suggestions
sheet.append(["Zeynep", "Kaya", "z.kaya@avzgroup.com", "Ali Veli Zeynep Ltd",
              "00000001", "Bounced"])
# row 3: matches an officer who has RESIGNED -> should also get suggestions
sheet.append(["Gone", "Old", "gone.old@avzgroup.com", "Ali Veli Zeynep Ltd",
              "00000001", "Bounced"])
workbook.save("feat_in.xlsx")

original = ED.CompaniesHouseClient


class Patched(original):
    def __init__(self, *a, **k):
        original.__init__(self, *a, **k)
        self._session.get = fake_get


ED.CompaniesHouseClient = Patched
os.environ["CH_API_KEY"] = "fake"
# --plain keeps the input row order, so the rows can be checked positionally
ED.cli(["triage", "-i", "feat_in.xlsx", "-o", "feat_out.xlsx", "--plain"])
ED.CompaniesHouseClient = original

rows = list(openpyxl.load_workbook("feat_out.xlsx").active.iter_rows(values_only=True))
header, data = list(rows[0]), [list(r) for r in rows[1:]]
os.remove("feat_in.xlsx")
os.remove("feat_out.xlsx")

print("     columns: %s" % ", ".join(str(h) for h in header[6:]))
for name in ("companyhouse_names", "active_officer_suggestions"):
    check("column %r present" % name, name in header, True)

names_index = header.index("companyhouse_names")
sugg_index = header.index("active_officer_suggestions")
result_index = header.index("result")

check("current name only, no former names", data[0][names_index],
      "AVZ TRADING LIMITED")
check("active match -> no suggestions", data[0][sugg_index] or "", "")
check("active match result", (data[0][result_index] or "").split(":")[0], ED.R.ACTIVE)

print()
print("     -- unmatched contact --")
check("result", data[1][result_index], ED.R.NO_OFFICER)
check("suggestions capped at %d + remainder" % ED.ACTIVE_SUGGESTION_LIMIT,
      data[1][sugg_index],
      "John Smith (director) | Priya Patel (secretary) | +1 more")
check("resigned officer excluded from suggestions",
      "Gone Old" in (data[1][sugg_index] or ""), False)

print()
print("     -- contact matched but RESIGNED --")
check("result", (data[2][result_index] or "").split(":")[0], ED.R.RESIGNED)
check("resigned row also gets suggestions", data[2][sugg_index],
      "John Smith (director) | Priya Patel (secretary) | +1 more")
check("limit constant is 2", ED.ACTIVE_SUGGESTION_LIMIT, 2)

print()
print("=" * 84)
if failures:
    print("%d FAILURE(S):" % len(failures))
    for item in failures:
        print("   -", item)
    sys.exit(1)
print("FEATURE TESTS PASSED")
