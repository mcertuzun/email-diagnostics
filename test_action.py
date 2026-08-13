# -*- coding: utf-8 -*-
"""
Tests for the action column.

The point of interest is that the classification reads three things, not two.
detect_email_typo reports result_reason as generic_mailbox regardless of what
the domain turned out to be, so info@acme.co.uk and info@totallyunrelated.com
are identical by reason. The recorded domain verdict is what separates them.

No network access.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
os.chdir(HERE)

import openpyxl
import email_diagnostics as ED


class Resp(object):
    def __init__(self, payload):
        self._payload, self.status_code, self.headers = payload, 200, {}

    def json(self):
        return self._payload


ACTIVE = [{"name": "SMITH, John",
           "name_elements": {"forename": "John", "surname": "Smith"},
           "officer_role": "director"},
          {"name": "PATEL, Priya",
           "name_elements": {"forename": "Priya", "surname": "Patel"},
           "officer_role": "secretary"}]
RESIGNED = [{"name": "OLD, Gone",
             "name_elements": {"forename": "Gone", "surname": "Old"},
             "officer_role": "director", "resigned_on": "2021-05-01"}]
SHARED_SURNAME = [{"name": "BROWN, Alan",
                   "name_elements": {"forename": "Alan", "surname": "Brown"}},
                  {"name": "BROWN, Barry",
                   "name_elements": {"forename": "Barry", "surname": "Brown"}}]

REGISTRY = {"00000001": ACTIVE, "00000002": RESIGNED, "00000003": SHARED_SURNAME}
DISSOLVED = {"00000004"}


def fake_get(url, params=None, timeout=None):
    if url.endswith("officers"):
        number = url.rstrip("/").split("/")[-2]
        items = REGISTRY.get(number, [])
        return Resp({"items": items, "total_results": len(items),
                     "items_per_page": 100, "start_index": 0})
    number = url.split("/")[-1]
    status = "dissolved" if number in DISSOLVED else "active"
    return Resp({"company_name": "ACME TRADING LIMITED", "company_status": status})


# first, last, email, regnum, expected action, note
CASES = [
    ("John", "Smith", "john.smith@acme.co.uk", "00000001",
     ED.A.ALL_CORRECT, "active, name matches, company domain"),
    ("John", "Smith", "info@acme.co.uk", "00000001",
     ED.A.ALL_CORRECT, "generic mailbox but the domain IS the company"),
    ("John", "Smith", "john.smith@gmail.com", "00000001",
     ED.A.NON_COMPANY_DOMAIN, "active, personal provider"),
    ("John", "Smith", "info@gmail.com", "00000001",
     ED.A.NON_COMPANY_DOMAIN, "generic mailbox on a personal provider"),
    ("John", "Smith", "info@totallyunrelated.com", "00000001",
     ED.A.MISMATCHED, "generic mailbox on an unrelated domain -> NOT all-correct"),
    ("John", "Smith", "john.smith@totallyunrelated.com", "00000001",
     ED.A.MISMATCHED, "active, domain unrelated to the company"),
    ("John", "Smith", "xq7z9@acme.co.uk", "00000001",
     ED.A.MISMATCHED, "active, address bears no resemblance to the name"),
    ("John", "Smith", "jhon.smith@acme.co.uk", "00000001",
     ED.A.FIX_ADDRESS, "typo you can correct"),
    ("John", "Smith", "not-an-email", "00000001",
     ED.A.FIX_ADDRESS, "malformed address"),
    ("Gone", "Old", "gone.old@acme.co.uk", "00000002",
     ED.A.FIND_NEW_CONTACT, "person has resigned"),
    ("Nobody", "Here", "nobody.here@acme.co.uk", "00000001",
     ED.A.FIND_NEW_CONTACT, "no officer by that name"),
    ("Someone", "Else", "someone@acme.co.uk", "00000004",
     ED.A.FIND_NEW_CONTACT, "company dissolved"),
    ("Chris", "Brown", "c.brown@acme.co.uk", "00000003",
     ED.A.UNCERTAIN_MATCH, "surname shared by two officers, forename does not match"),
    ("John", "Smith", "john.smith@acme.co.uk", "",
     ED.A.FIX_DATA, "regnum missing"),
]

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["first_name", "last_name", "email", "company", "regnum", "status"])
for first, last, email, regnum, _action, _note in CASES:
    sheet.append([first, last, email, "Acme Trading Ltd", regnum, "Bounced"])
workbook.save("act_in.xlsx")

original = ED.CompaniesHouseClient


class Patched(original):
    def __init__(self, *a, **k):
        original.__init__(self, *a, **k)
        self._session.get = fake_get


ED.CompaniesHouseClient = Patched
os.environ["CH_API_KEY"] = "fake"
ED.cli(["triage", "-i", "act_in.xlsx", "-o", "act_out.xlsx"])
ED.CompaniesHouseClient = original

rows = list(openpyxl.load_workbook("act_out.xlsx").active.iter_rows(values_only=True))
header, data = list(rows[0]), [list(r) for r in rows[1:]]
action_index = header.index("action")
result_index = header.index("result")
reason_index = header.index("result_reason")
os.remove("act_in.xlsx")
os.remove("act_out.xlsx")

problems = []
print("=" * 108)
print(" ACTION CLASSIFICATION")
print("=" * 108)
print("%-32s %-32s %-30s %s" % ("EXPECTED ACTION", "GOT", "result_reason", "OK?"))
print("-" * 108)
for case, row in zip(CASES, data):
    expected = case[4]
    got = row[action_index] or ""
    ok = got == expected
    if not ok:
        problems.append("%s -> expected %r, got %r  (%s)" % (case[2], expected, got, case[5]))
    print("%-32s %-32s %-30s %s" % (expected, got, row[reason_index] or "", "OK" if ok else "FAIL"))
    print("    %-28s %s" % (case[2][:28], case[5]))

print()
print("-" * 108)
print("action is the FIRST added column: %r" % header[6])
if header[6] != "action":
    problems.append("action is not the first added column: %r" % header[6])

print()
print("=" * 108)
if problems:
    print("%d FAILURE(S):" % len(problems))
    for item in problems:
        print("   -", item)
    sys.exit(1)
print("ACTION TESTS PASSED")
