# -*- coding: utf-8 -*-
"""End to end test: build a real .xlsx, run the pipeline, verify the output."""
import csv
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(HERE)

import openpyxl
import email_diagnostics as ED

# ------------------------------------------------------------------
# 1) Build a realistic input file, deliberately messy
# ------------------------------------------------------------------
SRC = os.path.join(HERE, "contacts_test.xlsx")
OUT = os.path.join(HERE, "out_test.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
# Deliberately messy headers: spaces, capitals, a BOM
ws.append(["﻿First Name ", "last_name", "Email", "company", "regnum", " status"])
ROWS = [
    ("Mr John",        "Smith",         "jhon.smith@acme.co.uk",  "Acme Trading Ltd",  "01234567", "Bounced"),
    ("Dr. Jane (Janie)", "Doe MBE",     "jane.doe@acme.co.uk",    "Acme Trading Ltd",  1234567,    "bounced back"),
    ("",               "Robert Brown",  "r.brown@acmee.co.uk",    "Acme Trading Ltd",  "01234567", "Delivery Failed"),
    ("Sarah",          "Jones",         "info@acme.co.uk",        "Acme Trading Ltd",  "SC123456", "hard bounce"),
    ("Peter",          "Nobody",        "",                       "Acme Trading Ltd",  "01234567", "undelivered"),
    ("Ali",            "Veli",          "not-an-email",           "Ali Veli Zeynep Ltd", "07654321", "failed"),
    ("Mehmet",         "Ozturk",        "m.ozturk@avz.co.uk",     "Ali Veli Zeynep Ltd", "07654321", "bounced"),
    ("Kate",           "Wilson",        "kate.wilson@acme.co.uk", "Acme Trading Ltd",  "01234567", "Delivered"),   # must be filtered out
    ("Tom",            "Baker",         "tom.baker@acme.co.uk",   "Acme Trading Ltd",  "01234567", "Blocked"),     # must be filtered out
]
for row in ROWS:
    ws.append(list(row))
wb.save(SRC)
print("Input built: %s (%d data rows)" % (os.path.basename(SRC), len(ROWS)))

# ------------------------------------------------------------------
# 2) Point the module at the test files and run
# ------------------------------------------------------------------
ED.INPUT_FILE = SRC
ED.OUTPUT_FILE = OUT
ED.DRY_RUN = True          # never call Companies House
ED.DEBUG = True            # include the audit columns
if os.path.exists(OUT):
    os.remove(OUT)

before_mtime = os.path.getmtime(SRC)
before_size = os.path.getsize(SRC)

print("-" * 78)
ED.main()
print("-" * 78)

# ------------------------------------------------------------------
# 3) Assertions
# ------------------------------------------------------------------
problems = []

# a) The input file must be untouched
if os.path.getmtime(SRC) != before_mtime or os.path.getsize(SRC) != before_size:
    problems.append("THE INPUT FILE WAS MODIFIED!")
else:
    print("OK   input file untouched")

# b) Exactly ONE output file
xlsx_files = sorted(f for f in os.listdir(HERE) if f.endswith(".xlsx"))
if xlsx_files != ["contacts_test.xlsx", "out_test.xlsx"]:
    problems.append("Unexpected files: %s" % xlsx_files)
else:
    print("OK   exactly one output file created")

# c) Read the output
wb2 = openpyxl.load_workbook(OUT)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))
header = list(rows[0])
data = [list(r) for r in rows[1:]]

print("OK   output columns: %s" % ", ".join(str(h) for h in header[:10]))

# d) Original columns preserved
if [str(h).strip() for h in header[:6]] != ["First Name", "last_name", "Email", "company", "regnum", "status"]:
    problems.append("Original headers not preserved: %r" % header[:6])
else:
    print("OK   original headers preserved")

# e) Filtering: 7 of 9 rows survive (Delivered and Blocked excluded)
if len(data) != 7:
    problems.append("Expected 7 rows, got %d" % len(data))
else:
    print("OK   status filter: 9 -> 7 rows (Delivered and Blocked excluded)")

result_idx = header.index("result")
reason_idx = header.index("result_reason")
regnum_dbg_idx = header.index("dbg_regnum_used")

expected = {
    "jhon.smith@acme.co.uk": ED.R.FIRST_NAME_TYPO,
    "not-an-email":          ED.R.MALFORMED_EMAIL,
    "":                      ED.R.MISSING_EMAIL,
    "r.brown@acmee.co.uk":   ED.R.DOMAIN_TYPO,
}
email_idx = 2
print()
print("%-26s %-24s %-30s %s" % ("EMAIL", "RESULT", "REASON", "REGNUM"))
print("-" * 100)
for row in data:
    email = row[email_idx] or ""
    print("%-26s %-24s %-30s %s" % (email[:26], str(row[result_idx])[:24],
                                    str(row[reason_idx])[:30], row[regnum_dbg_idx]))
    if email in expected and row[result_idx] != expected[email]:
        problems.append("%r -> expected %r, got %r" % (email, expected[email], row[result_idx]))

print()
# f) regnum zero padding: the numeric 1234567 must become 01234567
regnums = set(r[regnum_dbg_idx] for r in data)
if "1234567" in regnums:
    problems.append("regnum zero padding failed: %s" % regnums)
else:
    print("OK   regnum zero padding works: %s" % sorted(regnums))

# g) No row may be left without a result
if any(not r[result_idx] for r in data):
    problems.append("A row has an empty result")
else:
    print("OK   every row has a result")

# ------------------------------------------------------------------
# 4) CSV path and command line interface
#    A difficult Windows-style CSV: ';' delimiter, cp1254, accented text
# ------------------------------------------------------------------
print()
print("=" * 78)
print("CSV + CLI TEST")
print("=" * 78)

CSV_IN = os.path.join(HERE, "liste_test.csv")
CSV_OUT = os.path.join(HERE, "liste_test_sonuc.csv")
csv_rows = [
    "first_name;last_name;email;company;regnum;status",
    "Mr John;Smith;jhon.smith@acme.co.uk;Acme Trading Ltd;01234567;Bounced",
    "Mehmet;Öztürk;m.ozturk@avz.co.uk;Ali Veli Zeynep Ltd;07654321;bounced back",
    "Şükrü;Yılmaz;info@acme.co.uk;Acme Trading Ltd;01234567;hard bounce",
    "Kate;Wilson;kate.wilson@acme.co.uk;Acme Trading Ltd;01234567;Delivered",
]
io.open(CSV_IN, "w", encoding="cp1254", newline="").write("\r\n".join(csv_rows) + "\r\n")

exit_code = ED.cli(["triage", "--input", CSV_IN, "--output", CSV_OUT,
                    "--verbose", "--dry-run"])
if exit_code != 0:
    problems.append("CLI returned a non-zero exit code: %s" % exit_code)
else:
    print("OK   CLI exit code 0")

with io.open(CSV_OUT, "r", encoding="utf-8-sig", newline="") as handle:
    out_rows = list(csv.reader(handle, delimiter=";"))

if len(out_rows) != 4:              # 1 header + 3 bounced rows
    problems.append("CSV output rows: expected 4, got %d" % len(out_rows))
else:
    print("OK   ';' delimiter detected and preserved in the output")

if out_rows[0][:6] != ["first_name", "last_name", "email", "company", "regnum", "status"]:
    problems.append("CSV headers not preserved: %r" % out_rows[0][:6])
else:
    print("OK   CSV original columns preserved")

turkish = [r[1] for r in out_rows[1:]]
if "Öztürk" not in turkish or "Yılmaz" not in turkish:
    problems.append("Accented characters were mangled: %r" % turkish)
else:
    print("OK   accented characters survive cp1254 -> utf-8-sig")

csv_result_idx = out_rows[0].index("result")
if out_rows[1][csv_result_idx] != ED.R.FIRST_NAME_TYPO:
    problems.append("Wrong CSV diagnosis: %r" % out_rows[1][csv_result_idx])
else:
    print("OK   diagnosis correct on the CSV path")

os.remove(CSV_IN)
os.remove(CSV_OUT)

print()
print("=" * 78)
if problems:
    print("%d PROBLEM(S):" % len(problems))
    for p in problems:
        print("   -", p)
    sys.exit(1)
print("END TO END TEST PASSED")
