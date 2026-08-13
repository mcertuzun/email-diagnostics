# -*- coding: utf-8 -*-
"""
====================================================================
 EMAIL BOUNCE DIAGNOSTICS  -  Companies House assisted
====================================================================

Purpose:
    Analyse bounced email rows and write ONE "result" diagnosis per row.

Pipeline:
    1) Load the input file and validate the required columns
    2) Keep only rows whose status is in the bounce family
    3) Clean first_name / last_name / company / email for matching
    4) Look the company up on Companies House (parallel, rate limited)
    5) Check the email against the verified name
    6) Handle every failure explicitly
    7) Write ONE output file

HARD CONSTRAINT:
    This script never contacts a mail server. No SMTP handshake, no RCPT TO
    probe, no verification service, not even a DNS/MX lookup. The Companies
    House REST API is the only outbound connection.
    An address can therefore NEVER be proven valid; the result is an
    inference from its consistency with the name and the company.

RUNTIME:
    Written for Python 3.6.5 on Windows.
    - no dataclasses (3.7+)
    - no pandas / numpy: pandas turns regnum into a number and drops the
      leading zeros, which becomes a silent 404 from Companies House
    - no rapidfuzz / python-Levenshtein: the edit distance is implemented
      here, so no compiler is needed

REQUIREMENTS:
    pip install openpyxl==3.0.10 requests==2.27.1

API KEY (never written into the code):
    Windows, persistent :  setx CH_API_KEY "your_key"   then reopen the terminal
    Windows, session    :  set CH_API_KEY=your_key      (no quotes)
    PowerShell          :  $env:CH_API_KEY="your_key"
    macOS / Linux       :  export CH_API_KEY="your_key"

    A .env file next to this script works too, and needs no restart.
    Get a key at https://developer.company-information.service.gov.uk/
    by creating a "REST API" application.
"""

from __future__ import print_function

import argparse
import csv
import io
import json
import logging
import os
import re
import sys
import threading
import time
import unicodedata
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("ERROR: openpyxl is not installed.  ->  pip install openpyxl==3.0.10")

try:
    import requests
except ImportError:  # pragma: no cover
    sys.exit("ERROR: requests is not installed.  ->  pip install requests==2.27.1")


# ====================================================================
# SECTION 0 - SETTINGS  (the only part you normally need to edit)
# ====================================================================

# --- File paths ----------------------------------------------------
# These are DEFAULTS; --input / --output override them:
#   python email_diagnostics.py triage --input list.csv --output result.csv --verbose
INPUT_FILE = r"contacts.xlsx"                    # .xlsx / .xlsm / .csv / .tsv
OUTPUT_FILE = r"result_email_diagnostics.xlsx"   # the ONE output file
INPUT_SHEET = None                               # None = first sheet, or "Sheet1"
INPUT_DELIMITER = None                           # CSV delimiter; None = detect
INPUT_ENCODING = None                            # CSV encoding; None = detect

# --- Run mode ------------------------------------------------------
DEBUG = False          # True -> add the audit columns to the output
DRY_RUN = False        # True -> never call Companies House (offline)
MAX_ROWS = None        # e.g. 50 -> process only the first 50 bounced rows
MAX_COMPANIES = None   # e.g. 10 -> query at most 10 DISTINCT regnums

# LOOKUP_MODE:
#   "ch_first"   -> EVERY row goes to Companies House. The official name,
#                   middle names included, is fetched and the email is judged
#                   against it. N rows = N lookups. More reliable, costs more.
#   "typo_first" -> Typo check first; a row with a typo never reaches the API.
#                   Cheaper, but the name check relies on the raw input data.
LOOKUP_MODE = "ch_first"

# Fetch the company profile as well as the officer list. This is what fills
# the companyhouse_names column and detects dissolved companies. It costs one
# extra request per company; turn it off with --no-company-profile.
FETCH_COMPANY_PROFILE = True

# --- Required columns ----------------------------------------------
REQUIRED_COLUMNS = ["first_name", "last_name", "email", "company", "regnum", "status"]

# --- Status filtering ----------------------------------------------
# A row is ANALYSED if its status contains any of these.
PROBLEMATIC_STATUS_KEYWORDS = [
    "bounce", "bounced", "bounced back", "hard bounce", "soft bounce",
    "undelivered", "not delivered", "delivery failed", "delivery failure",
    "failed", "failure", "rejected",
]
# A row is EXCLUDED if its status contains any of these; this wins.
# "blocked" is excluded deliberately: it comes from spam filtering or IP
# reputation and says nothing about a wrong address or a departed person.
EXCLUDED_STATUS_KEYWORDS = ["blocked", "block", "spam", "unsubscribed", "suppressed"]

# --- Cleaning vocabularies -----------------------------------------
TITLES = {
    "mr", "mrs", "ms", "miss", "mx", "dr", "prof", "professor", "sir", "lord",
    "lady", "dame", "rev", "reverend", "capt", "captain", "hon", "madam", "master",
}
ROLE_WORDS = {
    "ceo", "cfo", "coo", "cto", "cio", "md", "director", "managing", "founder",
    "cofounder", "co-founder", "partner", "manager", "owner", "chairman",
    "chairwoman", "chair", "president", "vp", "head", "principal", "consultant",
    "secretary", "proprietor", "executive", "officer",
}
# UK post-nominals and honours; they turn up in first_name and last_name alike
POST_NOMINALS = {
    "obe", "mbe", "cbe", "kbe", "dbe", "bem", "jp", "dl",
    "fca", "aca", "acca", "fcca", "cima", "acma", "mrics", "frics", "mciob",
    "ceng", "mieee", "miet", "mba", "bsc", "msc", "ba", "ma", "beng", "meng",
    "phd", "dphil", "llb", "llm", "bcom", "cpa",
    "jr", "junior", "snr", "sr", "senior", "ii", "iii", "iv",
}
# Surname particles, kept as part of the surname
SURNAME_PARTICLES = {
    "van", "von", "de", "del", "della", "der", "den", "di", "da", "du", "dos",
    "la", "le", "les", "mac", "mc", "st", "saint", "ter", "ten", "af", "al",
}
COMPANY_STOPWORDS = {
    "limited", "ltd", "ltd.", "plc", "llp", "llc", "incorporated", "inc",
    "corporation", "corp", "co", "company", "group", "holdings", "holding",
    "services", "service", "solutions", "uk", "gb", "the", "and", "international",
    "trading", "enterprises", "enterprise", "partners", "associates",
}
# Words that companies commonly append to an acronym when registering a
# domain: "avz" -> "avzgroup.com", "avzltd.co.uk", "avzuk.com".
ACRONYM_DOMAIN_SUFFIXES = [
    "ltd", "limited", "plc", "llp", "llc", "inc", "co", "company", "corp",
    "group", "holdings", "holding", "services", "service", "solutions",
    "uk", "gb", "global", "international", "intl", "partners", "associates",
    "consulting", "consultancy", "trading", "enterprises", "online", "digital",
]

GENERIC_MAILBOXES = {
    "info", "information", "admin", "administrator", "office", "hello", "hi",
    "contact", "contactus", "enquiries", "enquiry", "inquiries", "inquiry",
    "sales", "support", "help", "helpdesk", "service", "customerservice",
    "accounts", "accounting", "account", "finance", "invoices", "invoice",
    "billing", "payments", "purchasing", "purchase", "orders", "order",
    "hr", "jobs", "careers", "recruitment", "marketing", "press", "media",
    "noreply", "no-reply", "donotreply", "mail", "email", "post", "reception",
    "team", "general", "main", "operations", "ops", "it", "webmaster", "director",
}
FREE_EMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "hotmail.co.uk", "outlook.com",
    "outlook.co.uk", "live.com", "live.co.uk", "msn.com", "yahoo.com",
    "yahoo.co.uk", "ymail.com", "aol.com", "icloud.com", "me.com", "mac.com",
    "btinternet.com", "sky.com", "virginmedia.com", "talktalk.net", "blueyonder.co.uk",
    "ntlworld.com", "protonmail.com", "proton.me", "gmx.com", "mail.com", "zoho.com",
}
# Multi-part TLDs, so the domain root is split correctly
MULTI_PART_TLDS = {
    "co.uk", "org.uk", "ltd.uk", "plc.uk", "me.uk", "net.uk", "sch.uk",
    "ac.uk", "gov.uk", "nhs.uk", "com.au", "co.nz", "co.za", "com.tr",
    "co.in", "com.sg", "co.jp", "com.br",
}
# Nickname groups, expanded both ways. Extend freely.
NICKNAME_GROUPS = [
    ["robert", "rob", "bob", "bobby", "robbie"],
    ["william", "will", "bill", "billy", "willy"],
    ["richard", "rich", "rick", "dick", "richie"],
    ["john", "jon", "jack", "johnny", "jonny"],
    ["james", "jim", "jimmy", "jamie"],
    ["thomas", "tom", "tommy"],
    ["michael", "mike", "mick", "micky", "mikey"],
    ["david", "dave", "davey"],
    ["stephen", "steven", "steve", "stevie"],
    ["christopher", "chris", "kit"],
    ["matthew", "matt", "matty"],
    ["nicholas", "nick", "nicky"],
    ["anthony", "tony", "ant"],
    ["edward", "ed", "eddie", "ted", "teddy", "ned"],
    ["samuel", "sam", "sammy"],
    ["benjamin", "ben", "benny"],
    ["andrew", "andy", "drew"],
    ["daniel", "dan", "danny"],
    ["joseph", "joe", "joey"],
    ["peter", "pete", "petey"],
    ["gregory", "greg"],
    ["charles", "charlie", "chas", "chuck"],
    ["frederick", "fred", "freddie"],
    ["henry", "harry", "hal", "harold"],
    ["alexander", "alexandra", "alex", "sandy", "lex"],
    ["patrick", "pat", "paddy"],
    ["timothy", "tim", "timmy"],
    ["ronald", "ron", "ronnie"],
    ["donald", "don", "donnie"],
    ["kenneth", "ken", "kenny"],
    ["lawrence", "laurence", "larry", "laurie"],
    ["philip", "phillip", "phil"],
    ["raymond", "ray"],
    ["terence", "terry"],
    ["albert", "al", "bert", "bertie"],
    ["arthur", "art", "artie"],
    ["george", "geordie"],
    ["katherine", "catherine", "kathryn", "kate", "katie", "kathy", "cathy", "kay"],
    ["elizabeth", "liz", "lizzie", "beth", "betty", "eliza", "libby", "betsy"],
    ["susan", "sue", "susie", "suzy"],
    ["jennifer", "jen", "jenny"],
    ["margaret", "maggie", "peggy", "meg", "greta"],
    ["patricia", "pat", "patty", "trish", "tricia"],
    ["rebecca", "becky", "becca"],
    ["victoria", "vicky", "vicki", "tori"],
    ["christine", "christina", "chris", "chrissy", "tina"],
    ["deborah", "debbie", "deb", "debra"],
    ["sandra", "sandy"],
    ["barbara", "barb", "babs"],
    ["joanne", "joanna", "jo", "joan"],
    ["pamela", "pam"],
    ["angela", "angie"],
    ["theresa", "teresa", "terri", "tess"],
    ["cynthia", "cindy"],
    ["dorothy", "dot", "dottie"],
    ["frances", "francis", "fran", "frankie"],
    ["gillian", "gill"],
    ["helen", "ellie", "nell"],
    ["jacqueline", "jackie", "jacqui"],
    ["kimberley", "kim"],
    ["michelle", "shelly"],
    ["nicola", "nicky", "nikki"],
    ["samantha", "sam", "sammy"],
    ["sarah", "sara", "sally"],
    ["stephanie", "steph"],
    ["abigail", "abby"],
    ["amanda", "mandy"],
    ["caroline", "carolyn", "carrie", "caz"],
    ["eleanor", "ellie", "nora"],
    ["isabella", "isabel", "izzy", "bella"],
]

# --- Thresholds ----------------------------------------------------
TYPO_MAX_DISTANCE_SHORT = 1   # allowed distance when the local part is <= 8 chars
TYPO_MAX_DISTANCE_LONG = 2    # allowed distance when the local part is longer
MIN_LOCAL_LEN_FOR_TYPO = 5    # never call a shorter local part a typo
DOMAIN_TYPO_MAX_DISTANCE = 2  # allowed distance on the domain root
MIN_DOMAIN_LEN_FOR_TYPO = 4   # never call a shorter domain root a typo
SURNAME_MAX_DISTANCE = 1      # allowed distance when matching an officer surname
ACTIVE_SUGGESTION_LIMIT = 5   # max names listed in active_officer_suggestions

# --- Companies House -----------------------------------------------
CH_API_BASE = "https://api.company-information.service.gov.uk"
CH_API_KEY_ENV = "CH_API_KEY"     # environment variable name
CH_TIMEOUT = 20                   # seconds
CH_MAX_RETRIES = 3                # for 429 / 5xx / timeout
CH_RATE_LIMIT_PER_SEC = 1.8       # documented limit: 600 requests / 5 min = 2.0/s
CH_WORKERS = 4                    # parallel threads
CH_PAGE_SIZE = 100                # officers page size (a CEILING, not a guarantee)
CH_MAX_OFFICERS = 2000            # most officers scanned for one company
CH_MAX_REQUESTS = None            # cap on total HTTP requests (None = no cap)
CH_CA_BUNDLE = None               # path to a corporate root certificate (.pem)
CH_VERIFY_SSL = True              # False = certificate verification OFF (unsafe)

# --- Logging -------------------------------------------------------
LOG_LEVEL = logging.INFO

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("diagnostics")

# Version: check with 'python email_diagnostics.py --version'.
# Running a stale copy is the most common source of confusion on Windows.
__version__ = "2.0.0"

# Command line defaults are captured ONCE, here. Reading the module globals
# when the parser is built would let one run's settings leak into the next,
# because apply_cli_args writes back to those same globals. That only shows
# up when cli() runs more than once in a process, which is exactly what the
# tests do -- and it silently changed how many companies were queried.
_DEFAULTS = {}


def _capture_defaults():
    _DEFAULTS.update({
        "input": INPUT_FILE, "output": OUTPUT_FILE, "sheet": INPUT_SHEET,
        "limit": MAX_ROWS, "limit_companies": MAX_COMPANIES,
        "mode": LOOKUP_MODE, "company_profile": FETCH_COMPANY_PROFILE,
        "workers": CH_WORKERS, "rate": CH_RATE_LIMIT_PER_SEC,
        "max_requests": CH_MAX_REQUESTS, "ca_bundle": CH_CA_BUNDLE,
    })


# ====================================================================
# SECTION 1 - FIXED RESULT VALUES
# ====================================================================

_capture_defaults()


class R(object):
    """The fixed set of values the result column can take."""
    MISSING_EMAIL = "missing_email"
    MALFORMED_EMAIL = "malformed_email"
    FIRST_NAME_TYPO = "first_name_typo"
    SURNAME_TYPO = "surname_typo"
    BOTH_TYPO = "first_name_and_surname_typo"
    DOMAIN_TYPO = "domain_typo"
    MISSING_REGNUM = "missing_regnum"
    COMPANY_NOT_FOUND = "company_not_found"
    CH_SKIPPED = "companies_house_skipped"
    COMPANY_DISSOLVED = "company_dissolved"
    LOOKUP_FAILED = "companies_house_lookup_failed"
    NO_OFFICER = "no_officer_match_found"
    ACTIVE = "active_officer_match"
    RESIGNED = "resigned_officer_match"
    POSSIBLE_ACTIVE = "possible_officer_match_active"
    POSSIBLE_RESIGNED = "possible_officer_match_resigned"


class RSN(object):
    """The fixed set of values the result_reason column can take."""
    PATTERN_OK = "email_matches_expected_pattern"
    GENERIC = "generic_mailbox"
    PERSONAL_DOMAIN = "personal_email_domain"
    UNRECOGNISED = "email_pattern_unrecognised"
    DOMAIN_NOT_MATCHED = "domain_not_matched"
    CLOSE_TO_PATTERN = "close_to_expected_pattern"
    WITH_MIDDLE_NAME = "matched_including_middle_name"
    SURNAME_ONLY = "surname_only_match"
    MULTIPLE_OFFICERS = "multiple_possible_officers"
    API_ERROR = "api_error"
    NO_NAME = "name_fields_empty"


# ====================================================================
# SECTION 2 - GENERAL HELPERS
# ====================================================================

def strip_accents(text):
    """Fold accents: Sukru -> sukru, Ozturk -> ozturk, O'Brien -> o'brien."""
    if not text:
        return ""
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def normalize(text):
    """Normalise for comparison: no accents, lower case, single spaces."""
    if text is None:
        return ""
    text = strip_accents(str(text))
    text = text.replace("﻿", " ").replace(" ", " ")
    text = text.lower().strip()
    return re.sub(r"\s+", " ", text)


def alnum_only(text):
    """Keep letters and digits only: 'john.smith' -> 'johnsmith'."""
    return re.sub(r"[^a-z0-9]", "", normalize(text))


def letters_only(text):
    """Keep letters only, dropping digits as well."""
    return re.sub(r"[^a-z]", "", normalize(text))


def edit_distance(a, b, max_distance=None):
    """
    Damerau-Levenshtein (optimal string alignment) distance, in pure
    Python so that no compiler or extra package is needed.

    Why not plain Levenshtein:
        The most common typing mistake is two letters swapping places
        ('jhon' for 'john', 'smiht' for 'smith'). Plain Levenshtein scores
        that as 2 edits, which pushes it past the threshold; Damerau
        scores it as 1.

    Passing max_distance lets it bail out early once that is exceeded.
    """
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    if max_distance is not None and abs(len(a) - len(b)) > max_distance:
        return max_distance + 1

    len_b = len(b)
    before_previous = None
    previous = list(range(len_b + 1))

    for i in range(1, len(a) + 1):
        current = [i] + [0] * len_b
        row_min = i
        for j in range(1, len_b + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            value = min(previous[j] + 1,          # deletion
                        current[j - 1] + 1,       # insertion
                        previous[j - 1] + cost)   # substitution
            # transposition
            if (i > 1 and j > 1
                    and a[i - 1] == b[j - 2]
                    and a[i - 2] == b[j - 1]):
                value = min(value, before_previous[j - 2] + cost)
            current[j] = value
            if value < row_min:
                row_min = value
        if max_distance is not None and row_min > max_distance:
            return max_distance + 1
        before_previous = previous
        previous = current
    return previous[len_b]


def build_nickname_map():
    """['robert','bob'] -> {'robert': {'robert','bob'}, 'bob': {'robert','bob'}}"""
    mapping = {}
    for group in NICKNAME_GROUPS:
        members = set(normalize(name) for name in group if name)
        for member in members:
            mapping.setdefault(member, set()).update(members)
    return mapping


NICKNAME_MAP = build_nickname_map()


def expand_nicknames(name):
    """Every nickname and formal variant of a given name."""
    key = normalize(name)
    if not key:
        return set()
    return set(NICKNAME_MAP.get(key, {key}))


def cell_to_text(value):
    """
    Convert an openpyxl cell value to text safely.
    Turns floats like 1234567.0 into 1234567, which matters for regnum.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, int):
        return str(value)
    return str(value).replace("﻿", "").strip()


# ====================================================================
# SECTION 3 - LOADING AND COLUMN VALIDATION
# ====================================================================

def normalize_header(header):
    """'  First Name ' -> 'first_name'. Strips BOM and non-breaking spaces."""
    text = (header or "")
    text = text.replace("﻿", "").replace(" ", " ")
    text = strip_accents(text).strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    text = re.sub(r"[^a-z0-9_]", "", text)
    return re.sub(r"_+", "_", text).strip("_")


def is_csv_path(path):
    """Decide from the extension whether this is CSV/TSV or Excel."""
    return os.path.splitext(path)[1].lower() in (".csv", ".tsv", ".txt")


def _sniff_delimiter(header_line):
    """
    Guess the delimiter from the header row.
    European Excel exports use ';', so assuming ',' is not good enough.
    """
    counts = [(header_line.count(sep), sep) for sep in [";", ",", "\t", "|"]]
    counts.sort(reverse=True)
    return counts[0][1] if counts[0][0] > 0 else ","


def _read_text_lines(path, encoding=None):
    """
    Read the CSV with the right encoding.
    Windows exports are commonly utf-8-sig (with BOM), cp1254 or cp1252.
    """
    encodings = [encoding] if encoding else ["utf-8-sig", "utf-8", "cp1254", "cp1252", "latin-1"]
    last_error = None
    for candidate in encodings:
        try:
            with io.open(path, "r", encoding=candidate, newline="") as handle:
                content = handle.read()
            if candidate != encodings[0]:
                log.warning("Detected encoding '%s'.", candidate)
            return content, candidate
        except (UnicodeDecodeError, LookupError) as exc:
            last_error = exc
            continue
    raise ValueError("Could not decode the file: {}  ({})\n"
                     "Set it explicitly with --encoding.".format(path, last_error))


def _load_csv(path, delimiter=None, encoding=None):
    """Read CSV/TSV. Everything stays text, so regnum keeps its leading zeros."""
    content, used_encoding = _read_text_lines(path, encoding)
    if not content.strip():
        raise ValueError("The CSV file is empty: {}".format(path))

    first_line = content.split("\n", 1)[0]
    sep = delimiter or _sniff_delimiter(first_line)
    log.info("Reading CSV (delimiter=%r, encoding=%s)", sep, used_encoding)

    reader = csv.reader(io.StringIO(content), delimiter=sep)
    try:
        header_row = next(reader)
    except StopIteration:
        raise ValueError("The CSV file is empty: {}".format(path))

    original_headers = [cell_to_text(cell) for cell in header_row]
    while original_headers and not original_headers[-1]:
        original_headers.pop()
    if not original_headers:
        raise ValueError("The header row is empty.")
    width = len(original_headers)

    rows = []
    for raw_row in reader:
        values = [cell_to_text(cell) for cell in raw_row[:width]]
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        if not any(values):
            continue
        rows.append(values)

    return original_headers, [normalize_header(h) for h in original_headers], rows, sep, used_encoding


def load_data(path, sheet_name=None, delimiter=None, encoding=None):
    """
    Read the input file, choosing CSV/TSV or Excel by extension.

    Deliberately avoids pandas: every value is read as text, so regnum keeps
    its leading zeros and date-like fields are not reinterpreted.

    Returns (original_headers, normalized_headers, rows, meta) where each row
    is a list of strings the same length as the header, and meta carries the
    CSV delimiter and encoding (both None for Excel).
    """
    if not os.path.isfile(path):
        raise IOError("Input file not found: {}".format(os.path.abspath(path)))

    if is_csv_path(path):
        headers, normalized, rows, sep, enc = _load_csv(path, delimiter, encoding)
        log.info("Loaded %s rows, %s columns  (%s)",
                 len(rows), len(headers), os.path.basename(path))
        return headers, normalized, rows, {"delimiter": sep, "encoding": enc}

    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError(
            "Unsupported extension: {}\n"
            "Readable: .xlsx, .xlsm, .csv, .tsv, .txt  (.xls is not supported)".format(path))

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("Sheet not found: '{}'. Available sheets: {}"
                                 .format(sheet_name, ", ".join(workbook.sheetnames)))
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ValueError("The Excel file is empty.")

        original_headers = [cell_to_text(cell) for cell in header_row]
        # Trim trailing empty columns
        while original_headers and not original_headers[-1]:
            original_headers.pop()
        width = len(original_headers)
        if width == 0:
            raise ValueError("The header row is empty.")

        normalized_headers = [normalize_header(h) for h in original_headers]

        rows = []
        for raw_row in row_iter:
            values = [cell_to_text(cell) for cell in raw_row[:width]]
            if len(values) < width:
                values.extend([""] * (width - len(values)))
            if not any(values):      # skip fully empty rows
                continue
            rows.append(values)
    finally:
        workbook.close()

    log.info("Loaded %s rows, %s columns  (%s)", len(rows), width, os.path.basename(path))
    return original_headers, normalized_headers, rows, {"delimiter": None, "encoding": None}


def validate_columns(normalized_headers, required=REQUIRED_COLUMNS):
    """
    Check that every required column is present.
    Returns {column_name: index}; raises a clear error listing what is missing.
    """
    index_map = {}
    for position, name in enumerate(normalized_headers):
        if name and name not in index_map:
            index_map[name] = position

    missing = [name for name in required if name not in index_map]
    if missing:
        raise ValueError(
            "Missing required column(s): {}\nFound in the file: {}".format(
                ", ".join(missing),
                ", ".join(h for h in normalized_headers if h) or "(none)"
            )
        )
    return index_map


# ====================================================================
# SECTION 4 - STATUS FILTERING
# ====================================================================

def is_problematic_status(status_text):
    """
    Status matching that tolerates case and small wording differences.
    The exclusion list (blocked and friends) is checked first.
    """
    text = normalize(status_text)
    if not text:
        return False
    flat = re.sub(r"[^a-z ]", " ", text)
    tokens = set(flat.split())

    for keyword in EXCLUDED_STATUS_KEYWORDS:
        key = normalize(keyword)
        if key in tokens or key in flat:
            return False

    for keyword in PROBLEMATIC_STATUS_KEYWORDS:
        if normalize(keyword) in flat:
            return True
    return False


def filter_problematic_statuses(rows, status_index):
    """Keep only the rows whose status is in the bounce family."""
    kept = [row for row in rows if is_problematic_status(row[status_index])]
    log.info("Status filter: %s of %s rows are bounces.", len(kept), len(rows))
    return kept


# ====================================================================
# SECTION 5 - DATA CLEANING
# ====================================================================

_NOISE_WORDS = TITLES | ROLE_WORDS | POST_NOMINALS


def _tokenize_person_field(raw):
    """Split a name field into tokens, returning bracketed nicknames separately."""
    text = (raw or "").replace("﻿", " ")
    nicknames = [normalize(m) for m in re.findall(r"[\(\[\"']([^\)\]\"']+)[\)\]\"']", text)]
    text = re.sub(r"[\(\[\"'][^\)\]\"']*[\)\]\"']", " ", text)   # drop bracketed parts
    text = text.replace(",", " , ")
    tokens = [t for t in re.split(r"\s+", strip_accents(text).strip()) if t]
    return tokens, [n for n in nicknames if n]


def _is_noise_token(token):
    """Is this token a title, a role word or a post-nominal?"""
    key = re.sub(r"[^a-z]", "", normalize(token))
    return bool(key) and key in _NOISE_WORDS


def _clean_name_tokens(tokens):
    """
    Drop titles, role words, post-nominals and punctuation leftovers.
    Returns (real name parts, initials).
    """
    words = []
    initials = []
    for token in tokens:
        if token == ",":
            words.append(",")
            continue
        if _is_noise_token(token):
            continue
        core = re.sub(r"[^a-z\-']", "", normalize(token))
        if not core:
            continue
        if len(core) == 1:
            initials.append(core)
        else:
            words.append(core)
    return words, initials


def clean_first_name(raw):
    """
    Clean the first_name field.

    Handles:
      "Mr John Smith"    -> first='john', surname candidate 'smith'
      "Dr. Jane (Janie)" -> first='jane', nicknames=['janie']
      "John, CEO"        -> first='john'
      "J. Michael"       -> first='michael', initials=['j']

    Returns dict(first, middles, nicknames, initials).
    """
    tokens, nicknames = _tokenize_person_field(raw)
    words, initials = _clean_name_tokens(tokens)
    words = [w for w in words if w != ","]

    first = words[0] if words else ""
    middles = words[1:] if len(words) > 1 else []
    return {
        "first": first,
        "middles": middles,
        "nicknames": [n for n in nicknames if n and n != first],
        "initials": initials,
    }


def clean_last_name(raw):
    """
    Clean the last_name field.

    Handles:
      "Smith"        -> surname='smith'
      "John Smith"   -> surname='smith', extra_firsts=['john']
      "SMITH, John"  -> surname='smith', extra_firsts=['john']  (comma form)
      "van der Berg" -> surname='van der berg' (plus a 'vanderberg' variant)
      "Smith MBE"    -> surname='smith'
      "Mr Smith"     -> surname='smith'  (titles appear here too)

    Returns dict(surname, extra_firsts, initials).
    """
    tokens, _ = _tokenize_person_field(raw)
    words, initials = _clean_name_tokens(tokens)

    if "," in words:
        pivot = words.index(",")
        before = [w for w in words[:pivot] if w != ","]
        after = [w for w in words[pivot + 1:] if w != ","]
        # "SMITH, John" -> surname first
        if before:
            return {"surname": " ".join(before), "extra_firsts": after, "initials": initials}
        words = after

    words = [w for w in words if w != ","]
    if not words:
        return {"surname": "", "extra_firsts": [], "initials": initials}

    # Keep surname particles (van, de, mc...) attached to the surname
    surname_start = len(words) - 1
    while surname_start > 0 and words[surname_start - 1] in SURNAME_PARTICLES:
        surname_start -= 1

    surname = " ".join(words[surname_start:])
    extra_firsts = words[:surname_start]
    return {"surname": surname, "extra_firsts": extra_firsts, "initials": initials}


def resolve_person_name(raw_first, raw_last):
    """
    Read first_name and last_name together and produce the final
    forename / middle name / surname candidates.

    The cases that matter:
      - first_name empty, last_name "John Smith" -> first='john', surname='smith'
      - first_name "John Smith", last_name empty -> first='john', surname='smith'
      - both populated                           -> used as given
    """
    first_data = clean_first_name(raw_first)
    last_data = clean_last_name(raw_last)

    first = first_data["first"]
    middles = list(first_data["middles"])
    surname = last_data["surname"]
    extra_firsts = list(last_data["extra_firsts"])

    # If last_name held a full name ("John Smith"), the leading parts are forenames
    if extra_firsts:
        if not first:
            first = extra_firsts[0]
            middles.extend(extra_firsts[1:])
        else:
            middles.extend(extra_firsts)

    # If first_name held a full name and no surname was given
    if not surname and middles:
        surname = middles[-1]
        middles = middles[:-1]

    # Forename candidates: the name, its nicknames, and the nickname map
    first_candidates = set()
    for value in [first] + first_data["nicknames"]:
        if value:
            first_candidates.update(expand_nicknames(value))
    first_candidates = set(letters_only(c) for c in first_candidates if letters_only(c))

    # Surname variants: spaced, joined, hyphenated, and each hyphen part
    surname_variants = set()
    if surname:
        base = letters_only(surname)
        if base:
            surname_variants.add(base)
        surname_variants.add(re.sub(r"[^a-z\-]", "", normalize(surname)).replace("-", ""))
        for part in re.split(r"[\s\-]+", normalize(surname)):
            part = letters_only(part)
            if len(part) >= 3 and part not in SURNAME_PARTICLES:
                surname_variants.add(part)
    surname_variants = set(v for v in surname_variants if v)

    initials = set(first_data["initials"]) | set(last_data["initials"])

    return {
        "first": letters_only(first),
        "middles": [letters_only(m) for m in middles if letters_only(m)],
        "surname": letters_only(surname.replace(" ", "")),
        "surname_display": surname,
        "first_candidates": first_candidates,
        "surname_variants": surname_variants,
        "nicknames": first_data["nicknames"],
        "initials": initials,
    }


def clean_company_name(raw):
    """
    Strip legal suffixes and stop words from a company name.
    "Ali Veli Zeynep Trading Ltd." -> ['ali','veli','zeynep']

    The original value is never modified; this is only for matching.
    """
    text = normalize(raw)
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = [t for t in text.split() if t]
    meaningful = [t for t in tokens if t not in COMPANY_STOPWORDS and len(t) > 1]
    return meaningful if meaningful else tokens


class DomainCandidates(object):
    """
    Two sets of plausible domain roots for one company.

    exact  - every candidate, used for exact and containment matching
    fuzzy  - only the candidates that are safe to compare by edit distance

    The split matters. An acronym is information dense: 'avzgroup' and
    'xyzgroup' are two characters apart but are unrelated companies, so
    letting acronym forms into the distance comparison turns any similar
    acronym into a reported domain typo. Acronyms therefore go in exact
    only, while full-word forms like 'alivelizeynep' are long and
    distinctive enough for distance comparison to be meaningful.
    """

    __slots__ = ("exact", "fuzzy")

    def __init__(self, exact=None, fuzzy=None):
        self.exact = set(exact or ())
        self.fuzzy = set(fuzzy or ())

    def add(self, value, fuzzy_safe=True):
        if value:
            self.exact.add(value)
            if fuzzy_safe:
                self.fuzzy.add(value)

    def update(self, other):
        self.exact |= other.exact
        self.fuzzy |= other.fuzzy

    # Keep the object usable wherever a plain set was expected before.
    def __iter__(self):
        return iter(self.exact)

    def __len__(self):
        return len(self.exact)

    def __contains__(self, item):
        return item in self.exact


def generate_company_domain_candidates(tokens, extra_names=None):
    """
    Plausible domain roots for a company name.
    ['ali','veli','zeynep'] -> alivelizeynep, avz, avzgroup, ali, aliveli ...
    """
    candidates = DomainCandidates()
    if extra_names:
        for name in extra_names:
            candidates.update(generate_company_domain_candidates(clean_company_name(name)))

    if not tokens:
        return candidates

    joined = "".join(tokens)
    if joined:
        candidates.add(joined)
        candidates.add("-".join(tokens))

    # Acronym of the company name, and the acronym followed by a common
    # company word: "Ali Veli Zeynep Ltd" -> avz, avzltd, avzgroup, avzuk.
    # Only for names of TWO OR MORE words; a single-word company yields a
    # one-letter acronym that would match almost anything.
    # fuzzy_safe=False: see the DomainCandidates docstring.
    if len(tokens) >= 2:
        acronym = "".join(t[0] for t in tokens if t)
        if len(acronym) >= 2:
            candidates.add(acronym, fuzzy_safe=False)
            for suffix in ACRONYM_DOMAIN_SUFFIXES:
                candidates.add(acronym + suffix, fuzzy_safe=False)

    candidates.add(tokens[0])
    if len(tokens) >= 2:
        candidates.add(tokens[0] + tokens[1])
        candidates.add(tokens[0] + "-" + tokens[1])
        candidates.add(tokens[0] + tokens[1][0], fuzzy_safe=False)
        candidates.add(tokens[0][0] + tokens[1], fuzzy_safe=False)
    if len(tokens) >= 3:
        candidates.add(tokens[0] + tokens[1] + tokens[2])

    candidates.exact = set(c for c in candidates.exact if c and len(c) >= 2)
    candidates.fuzzy = set(c for c in candidates.fuzzy if c and len(c) >= 2)
    return candidates


EMAIL_RE = re.compile(r"^[a-z0-9!#$%&'*+/=?^_`{|}~\-]+(\.[a-z0-9!#$%&'*+/=?^_`{|}~\-]+)*"
                      r"@[a-z0-9]([a-z0-9\-]*[a-z0-9])?(\.[a-z0-9]([a-z0-9\-]*[a-z0-9])?)+$")


def split_domain(domain):
    """
    'mail.acme.co.uk' -> brand='acme', root='mail.acme'
    Multi-part TLDs (co.uk, org.uk ...) are stripped correctly.
    """
    parts = [p for p in domain.split(".") if p]
    if len(parts) < 2:
        return domain, domain
    tld_length = 2 if ".".join(parts[-2:]) in MULTI_PART_TLDS and len(parts) >= 3 else 1
    name_parts = parts[:-tld_length]
    if not name_parts:
        name_parts = [parts[0]]
    brand = name_parts[-1]
    return brand, ".".join(name_parts)


def parse_email(raw):
    """
    Lower-case the address, remove whitespace and validate its shape.
    Returns dict(ok, status, raw, email, local, domain, brand) where
    status is "ok", "missing" or "malformed".
    """
    text = (raw or "").strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("﻿", "").strip("<>").strip(";,")
    lowered = strip_accents(text).lower()

    if not lowered:
        return {"ok": False, "status": "missing", "raw": raw, "email": "",
                "local": "", "domain": "", "brand": ""}

    if lowered.count("@") != 1 or not EMAIL_RE.match(lowered) or ".." in lowered:
        return {"ok": False, "status": "malformed", "raw": raw, "email": lowered,
                "local": "", "domain": "", "brand": ""}

    local, domain = lowered.split("@", 1)
    brand, _root = split_domain(domain)
    return {"ok": True, "status": "ok", "raw": raw, "email": lowered,
            "local": local, "domain": domain, "brand": brand}


# ====================================================================
# SECTION 6 - EXPECTED EMAIL PATTERNS AND TYPO DETECTION
# ====================================================================

SEPARATORS = ["", ".", "_", "-"]


class Pattern(object):
    """One expected local-part pattern, plus which name part each piece came from."""

    __slots__ = ("text", "parts", "roles", "sep")

    def __init__(self, parts, roles, sep):
        self.parts = parts
        self.roles = roles
        self.sep = sep
        self.text = sep.join(parts)


def generate_expected_email_patterns(name_data):
    """
    Build the plausible local-part patterns from the forename, surname,
    middle names, nicknames and initials.

    For each separator ('', '.', '_', '-'):
        first.last, last.first, f.last, first.l, first.m.last,
        firstlast, lastfirst, flast, firstl, fmlast
    Without a separator:
        first, last, and the initial combinations fl, fm, fml
    """
    patterns = []
    firsts = set(f for f in name_data["first_candidates"] if f)
    surnames = set(s for s in name_data["surname_variants"] if s)
    middles = [m for m in name_data["middles"] if m]

    if not firsts and not surnames:
        return patterns

    def add(parts, roles, sep):
        parts = [p for p in parts if p]
        if len(parts) != len(roles) or not parts:
            return
        patterns.append(Pattern(parts, roles, sep))

    for surname in surnames:
        add([surname], ["surname"], "")
        for first in firsts:
            for sep in SEPARATORS:
                add([first, surname], ["first", "surname"], sep)
                add([surname, first], ["surname", "first"], sep)
                add([first[0], surname], ["first", "surname"], sep)
                add([first, surname[0]], ["first", "surname"], sep)
                add([surname, first[0]], ["surname", "first"], sep)
                for middle in middles:
                    add([first, middle, surname], ["first", "middle", "surname"], sep)
                    add([first, middle[0], surname], ["first", "middle", "surname"], sep)
                    add([first[0], middle[0], surname], ["first", "middle", "surname"], sep)

    for first in firsts:
        add([first], ["first"], "")

    # Initial combinations, for short addresses like "js@"
    # (Jasmine Susanne Smith -> js, jss, ss)
    initial_sources = []
    for first in firsts:
        initial_sources.append(first[0])
    middle_initials = [m[0] for m in middles]
    surname_initials = [s[0] for s in surnames]

    for fi in set(initial_sources):
        for si in set(surname_initials):
            add([fi, si], ["first", "surname"], "")
            add([si, fi], ["surname", "first"], "")
        for mi in set(middle_initials):
            add([fi, mi], ["first", "middle"], "")
            for si in set(surname_initials):
                add([fi, mi, si], ["first", "middle", "surname"], "")
    for extra in name_data["initials"]:
        for si in set(surname_initials):
            add([extra, si], ["first", "surname"], "")

    return patterns


def _strip_trailing_digits(text):
    """'john.smith2' -> ('john.smith', True)"""
    stripped = re.sub(r"\d+$", "", text)
    return (stripped, stripped != text) if stripped else (text, False)


def _score_pattern(local, pattern):
    """Edit distance between the local part and one expected pattern."""
    limit = TYPO_MAX_DISTANCE_LONG + 1
    return edit_distance(local, pattern.text, max_distance=limit)


def _attribute_difference(local, pattern):
    """
    Work out which name part the difference falls in.
    Returns a set containing 'first', 'surname', or both.
    """
    bad = set()
    if pattern.sep:
        local_parts = local.split(pattern.sep)
        if len(local_parts) == len(pattern.parts):
            for value, expected, role in zip(local_parts, pattern.parts, pattern.roles):
                if edit_distance(value, expected, max_distance=3) > 0:
                    bad.add("first" if role in ("first", "middle") else "surname")
            return bad

    # Concatenated pattern: search for the best split point
    if len(pattern.parts) == 2:
        expected_a, expected_b = pattern.parts
        best = None
        for cut in range(1, max(2, len(local))):
            left, right = local[:cut], local[cut:]
            total = (edit_distance(left, expected_a, max_distance=4)
                     + edit_distance(right, expected_b, max_distance=4))
            if best is None or total < best[0]:
                best = (total, edit_distance(left, expected_a, max_distance=4),
                        edit_distance(right, expected_b, max_distance=4))
        if best:
            _total, da, db = best
            if da > 0:
                bad.add("first" if pattern.roles[0] in ("first", "middle") else "surname")
            if db > 0:
                bad.add("first" if pattern.roles[1] in ("first", "middle") else "surname")
            return bad

    return {"first", "surname"}


def check_domain(email_info, company_candidates):
    """
    Compare the domain root against the company name candidates.

    The ORDER matters:
      1) exact match                    -> ok
      2) containment with a real affix  -> ok    ('acme' -> 'acmegroup', 'acmeuk')
      3) 1-2 characters apart           -> typo  ('acme' -> 'acmee')
      4) anything else                  -> unmatched (NOT a typo, continue to CH)

    Step 2 must come before step 3, otherwise 'acmegroup' looks like a typo.
    But the affix in step 2 must be at least 2 characters, otherwise 'acmee'
    (one letter too many) would be accepted as valid.

    Returns (verdict, reason) where verdict is
    "ok", "personal", "typo" or "unmatched".
    """
    domain = email_info["domain"]
    brand = email_info["brand"]

    if domain in FREE_EMAIL_DOMAINS:
        return "personal", RSN.PERSONAL_DOMAIN
    if not company_candidates:
        return "unmatched", RSN.DOMAIN_NOT_MATCHED

    brand_flat = re.sub(r"[^a-z0-9]", "", brand)
    if not brand_flat:
        return "unmatched", RSN.DOMAIN_NOT_MATCHED

    def flatten(values):
        return [f for f in (re.sub(r"[^a-z0-9]", "", c) for c in values) if f]

    flats = flatten(company_candidates)
    # Only these are compared by edit distance; acronym forms are excluded.
    fuzzy_flats = flatten(getattr(company_candidates, "fuzzy", company_candidates))

    # 1) Exact match
    if brand_flat in flats:
        return "ok", None

    # 2) Containment: the difference must be at least 2 characters
    for flat in flats:
        if len(flat) < MIN_DOMAIN_LEN_FOR_TYPO:
            continue     # keep short tokens like 'ali' from matching 'alibaba'
        if flat in brand_flat and len(brand_flat) - len(flat) >= 2:
            return "ok", None
        if brand_flat in flat and len(flat) - len(brand_flat) >= 2:
            return "ok", None

    # 3) Close but not equal -> typo
    if len(brand_flat) >= MIN_DOMAIN_LEN_FOR_TYPO:
        best = None
        for flat in fuzzy_flats:
            if len(flat) < MIN_DOMAIN_LEN_FOR_TYPO:
                continue
            distance = edit_distance(brand_flat, flat,
                                     max_distance=DOMAIN_TYPO_MAX_DISTANCE + 1)
            if best is None or distance < best:
                best = distance
        if best is not None and 0 < best <= DOMAIN_TYPO_MAX_DISTANCE:
            return "typo", RSN.CLOSE_TO_PATTERN

    # 4) Unrelated domain: NOT called a typo; it may be a group domain
    return "unmatched", RSN.DOMAIN_NOT_MATCHED


def detect_email_typo(email_info, name_data, company_candidates):
    """
    The email consistency check.

    Returns dict(terminal, result, reason, best_pattern, best_distance,
    domain_verdict). terminal=True means the row is decided here and
    Companies House is not consulted for it.

    Design rule: never call a WEAK difference a typo.
      - exact match, or only a separator difference -> consistent
      - 1-2 characters apart                        -> typo
      - nowhere near any pattern                    -> email_pattern_unrecognised
      - local part shorter than the minimum         -> never a typo (initials)
    """
    outcome = {"terminal": False, "result": None, "reason": None,
               "best_pattern": "", "best_distance": "", "domain_verdict": ""}

    local = email_info["local"]
    local_letters = re.sub(r"[^a-z0-9._\-]", "", local)

    domain_verdict, domain_reason = check_domain(email_info, company_candidates)
    outcome["domain_verdict"] = domain_verdict

    # --- Generic mailbox: not a person, so a typo verdict is meaningless ---
    if re.sub(r"[^a-z]", "", local) in GENERIC_MAILBOXES or local in GENERIC_MAILBOXES:
        if domain_verdict == "typo":
            outcome.update(terminal=True, result=R.DOMAIN_TYPO, reason=RSN.GENERIC)
        else:
            outcome["reason"] = RSN.GENERIC
        return outcome

    # --- Without a name there is nothing to compare against ---
    patterns = generate_expected_email_patterns(name_data)
    if not patterns:
        outcome["reason"] = RSN.NO_NAME
        if domain_verdict == "typo":
            outcome.update(terminal=True, result=R.DOMAIN_TYPO,
                           reason=RSN.CLOSE_TO_PATTERN)
        return outcome

    compare_local, _had_digits = _strip_trailing_digits(local_letters)
    compare_flat = alnum_only(compare_local)

    # 1) Exact match
    for pattern in patterns:
        if compare_local == pattern.text:
            outcome["best_pattern"] = pattern.text
            outcome["best_distance"] = 0
            outcome["reason"] = RSN.PATTERN_OK
            if domain_verdict == "typo":
                outcome.update(terminal=True, result=R.DOMAIN_TYPO,
                               reason=RSN.CLOSE_TO_PATTERN)
            elif domain_verdict == "unmatched":
                outcome["reason"] = RSN.DOMAIN_NOT_MATCHED
            elif domain_verdict == "personal":
                outcome["reason"] = RSN.PERSONAL_DOMAIN
            return outcome

    # 2) Separator-only difference (john.smith vs johnsmith) -> NOT a typo
    for pattern in patterns:
        if compare_flat and compare_flat == alnum_only(pattern.text):
            outcome["best_pattern"] = pattern.text
            outcome["best_distance"] = 0
            outcome["reason"] = RSN.PATTERN_OK
            if domain_verdict == "typo":
                outcome.update(terminal=True, result=R.DOMAIN_TYPO,
                               reason=RSN.CLOSE_TO_PATTERN)
            elif domain_verdict == "unmatched":
                outcome["reason"] = RSN.DOMAIN_NOT_MATCHED
            elif domain_verdict == "personal":
                outcome["reason"] = RSN.PERSONAL_DOMAIN
            return outcome

    # 3) Nearest pattern
    best_pattern = None
    best_distance = None
    for pattern in patterns:
        distance = _score_pattern(compare_local, pattern)
        if best_distance is None or distance < best_distance:
            best_distance = distance
            best_pattern = pattern
            if distance == 0:
                break

    outcome["best_pattern"] = best_pattern.text if best_pattern else ""
    outcome["best_distance"] = best_distance if best_distance is not None else ""

    threshold = (TYPO_MAX_DISTANCE_SHORT if len(compare_local) <= 8
                 else TYPO_MAX_DISTANCE_LONG)

    too_short = len(compare_local) < MIN_LOCAL_LEN_FOR_TYPO
    is_typo = (best_pattern is not None
               and best_distance is not None
               and 0 < best_distance <= threshold
               and not too_short)

    if is_typo:
        bad_parts = _attribute_difference(compare_local, best_pattern)
        if bad_parts == {"first"}:
            result = R.FIRST_NAME_TYPO
        elif bad_parts == {"surname"}:
            result = R.SURNAME_TYPO
        else:
            result = R.BOTH_TYPO
        outcome.update(terminal=True, result=result, reason=RSN.CLOSE_TO_PATTERN)
        return outcome

    # 4) Not close to any pattern: not a typo, continue to Companies House
    if domain_verdict == "typo":
        outcome.update(terminal=True, result=R.DOMAIN_TYPO, reason=RSN.UNRECOGNISED)
        return outcome

    if domain_verdict == "personal":
        outcome["reason"] = RSN.PERSONAL_DOMAIN
    elif domain_verdict == "unmatched":
        outcome["reason"] = RSN.DOMAIN_NOT_MATCHED
    else:
        outcome["reason"] = RSN.UNRECOGNISED
    return outcome


# ====================================================================
# SECTION 7 - COMPANIES HOUSE CLIENT
# ====================================================================

class CompaniesHouseAuthError(Exception):
    """401/403 - the key was rejected. The whole run stops."""


class CompanyNotFound(Exception):
    """404 - no company with that number."""


class LookupFailed(Exception):
    """Still failing after the retries."""


class SSLHandshakeError(Exception):
    """TLS handshake failed, usually a corporate proxy or antivirus intercepting."""


def ssl_help(exc):
    """Actionable guidance for an SSL failure."""
    return "\n".join([
        "TLS handshake failed (SSL bad handshake).",
        "",
        "Cause: {}".format(exc),
        "",
        "This is almost always TLS INTERCEPTION on the network. A corporate",
        "firewall (Zscaler, Netskope, Fortinet, Cisco Umbrella) or antivirus",
        "SSL scanning (Kaspersky, ESET, Avast) presents its own certificate,",
        "which is not in Python's certificate store.",
        "",
        "FIXES, in order:",
        "",
        "1) Use the Windows certificate store - cleanest on a corporate network:",
        "     pip install pip-system-certs",
        "   Your organisation's root certificate is already registered with",
        "   Windows; this package points requests at that store. No code change.",
        "",
        "2) Supply the root certificate as a file:",
        "     python email_diagnostics.py triage ... --ca-bundle C:\\path\\corp-root.pem",
        "   or via the environment: set REQUESTS_CA_BUNDLE=C:\\path\\corp-root.pem",
        "   (ask IT for the .pem, or export it on Windows from",
        "    certmgr.msc > Trusted Root Certification Authorities)",
        "",
        "3) If you must go through a proxy:",
        "     set HTTPS_PROXY=http://proxy.company.local:8080",
        "",
        "4) If the certificate store is stale:",
        "     pip install --upgrade certifi",
        "",
        "LAST RESORT) --insecure turns certificate verification off.",
        "   Be aware this sends your API key over an unverified connection.",
        "   Use it to confirm the diagnosis, not as a permanent fix.",
    ])


class RateLimiter(object):
    """
    Rate limiter shared across threads.
    The Companies House limit is 600 requests per 5 minutes (2.0/second).
    """

    def __init__(self, rate_per_second):
        self._interval = 1.0 / float(rate_per_second)
        self._lock = threading.Lock()
        self._next_slot = 0.0

    def acquire(self):
        with self._lock:
            now = time.time()
            wait = self._next_slot - now
            if wait > 0:
                time.sleep(wait)
                now = time.time()
            self._next_slot = max(now, self._next_slot) + self._interval


class CompaniesHouseClient(object):
    """Companies House REST client: throttling, retries and pagination."""

    def __init__(self, api_key, rate_per_second=CH_RATE_LIMIT_PER_SEC,
                 max_requests=CH_MAX_REQUESTS):
        self._session = requests.Session()
        self._session.auth = (api_key, "")     # Basic auth: key as username, blank password
        self._session.headers.update({"Accept": "application/json"})

        # TLS verification: use the corporate root certificate when given.
        if CH_CA_BUNDLE:
            self._session.verify = CH_CA_BUNDLE
            log.info("TLS: using corporate root certificate -> %s", CH_CA_BUNDLE)
        elif not CH_VERIFY_SSL:
            self._session.verify = False
            log.warning("!" * 62)
            log.warning("--insecure IS ON: the TLS certificate is NOT verified.")
            log.warning("Your API key travels over an unverified connection.")
            log.warning("Use this to confirm a diagnosis, never as a fix.")
            log.warning("!" * 62)
            try:
                requests.packages.urllib3.disable_warnings()
            except Exception:
                pass
        self._limiter = RateLimiter(rate_per_second)
        self.last_status = {}                  # regnum -> last HTTP code, for DEBUG

        # Quota tracking: count the requests rather than estimating them.
        self._max_requests = max_requests
        self._stats_lock = threading.Lock()
        self.stats = {"requests": 0, "retries": 0, "rate_limited": 0,
                      "pages": 0, "failed": 0, "server_page_size": 0}

    def _bump(self, key, amount=1):
        with self._stats_lock:
            self.stats[key] += amount

    def _reserve_request(self):
        """
        Increment the request counter. Returns False once the ceiling is
        reached, and the request is then never sent.
        """
        with self._stats_lock:
            if self._max_requests is not None and self.stats["requests"] >= self._max_requests:
                return False
            self.stats["requests"] += 1
            return self.stats["requests"]

    def _request(self, path, params=None):
        url = CH_API_BASE + path
        last_error = "unknown error"

        for attempt in range(1, CH_MAX_RETRIES + 1):
            sequence = self._reserve_request()
            if sequence is False:
                raise LookupFailed(
                    "request ceiling reached ({}). Raise it with --max-requests."
                    .format(self._max_requests))
            if attempt > 1:
                self._bump("retries")

            self._limiter.acquire()
            log.debug("CH request #%s (attempt %s): %s %s",
                      sequence, attempt, path, params or "")
            try:
                response = self._session.get(url, params=params, timeout=CH_TIMEOUT)
            except requests.exceptions.SSLError as exc:
                # An SSL handshake failure is PERMANENT. Retrying only wastes
                # time and quota, so fail immediately with usable guidance.
                raise SSLHandshakeError(ssl_help(exc))
            except requests.exceptions.RequestException as exc:
                last_error = "baglanti: {}".format(exc)
                time.sleep(min(2 ** attempt, 10))
                continue

            code = response.status_code

            if code == 200:
                self._bump("pages")
                try:
                    return response.json(), code
                except ValueError:
                    last_error = "gecersiz JSON"
                    time.sleep(1)
                    continue

            if code in (401, 403):
                raise CompaniesHouseAuthError(
                    "Companies House rejected the credentials (HTTP {}). "
                    "Check the key in {}.".format(code, CH_API_KEY_ENV)
                )
            if code == 404:
                raise CompanyNotFound("HTTP 404")
            if code == 429:
                retry_after = response.headers.get("Retry-After")
                try:
                    delay = float(retry_after) if retry_after else 5.0
                except (TypeError, ValueError):
                    delay = 5.0
                self._bump("rate_limited")
                log.warning("Rate limited (429), waiting %.1f s...", delay)
                time.sleep(min(delay, 60))
                last_error = "HTTP 429"
                continue
            if 500 <= code < 600:
                last_error = "HTTP {}".format(code)
                time.sleep(min(2 ** attempt, 10))
                continue

            last_error = "HTTP {}".format(code)
            break

        self._bump("failed")
        raise LookupFailed(last_error)

    def get_companies_house_officers(self, company_number):
        """
        Every officer record for the company, following pagination.
        Companies with a long officer history do not fit in one page.
        """
        officers = []
        start_index = 0
        while True:
            params = {"items_per_page": CH_PAGE_SIZE, "start_index": start_index}
            data, code = self._request("/company/{}/officers".format(company_number), params)
            self.last_status[company_number] = code

            if not isinstance(data, dict):
                raise LookupFailed("unexpected response shape")
            items = data.get("items")
            if not isinstance(items, list):
                items = []
            officers.extend(items)

            total = data.get("total_results")
            try:
                total = int(total)
            except (TypeError, ValueError):
                total = len(officers)

            # The server may return FEWER records than asked for, because
            # items_per_page is a ceiling rather than a guarantee. Advance by
            # what actually arrived, not by what was requested: otherwise an
            # extra request is sent and the records in between are silently
            # skipped.
            # Measure the page size the server ACTUALLY used. The maximum
            # for items_per_page is not documented, so read it from the
            # response instead of assuming a value.
            served = data.get("items_per_page")
            try:
                served = int(served)
            except (TypeError, ValueError):
                served = len(items)
            with self._stats_lock:
                if served > self.stats["server_page_size"]:
                    self.stats["server_page_size"] = served

            received = len(items)
            if received == 0:
                break                       # guard against an endless loop
            start_index += received
            if len(officers) >= total or start_index >= CH_MAX_OFFICERS:
                break
        return officers

    def get_company_profile(self, company_number):
        """Official company name, status and any former names."""
        data, code = self._request("/company/{}".format(company_number))
        self.last_status[company_number] = code
        if not isinstance(data, dict):
            raise LookupFailed("beklenmeyen yanit yapisi")
        return data


def _clean_key_value(value):
    """
    Strip surrounding whitespace and quotes from the key.
    On Windows `set CH_API_KEY="abc"` puts the quotes INSIDE the value and
    the API then returns 401, which is a hard failure to diagnose.
    """
    value = (value or "").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("\"", "'"):
        value = value[1:-1].strip()
    return value


def read_env_file(path):
    """
    Minimal .env reader, no extra package required.
    Returns the KEY=VALUE pairs; lines starting with '#' are comments.
    """
    values = {}
    if not os.path.isfile(path):
        return values
    try:
        with io.open(path, "r", encoding="utf-8-sig") as handle:
            for line in handle:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _sep, value = line.partition("=")
                key = key.strip()
                if key.lower().startswith("export "):
                    key = key[7:].strip()
                values[key] = _clean_key_value(value)
    except (IOError, UnicodeDecodeError):
        pass
    return values


def get_api_key():
    """
    Look for the Companies House key, in order:
        1) the environment variable
        2) a .env file in the working directory
        3) a .env file next to this script

    Returns (key, where_it_came_from), or ("", "") if nothing was found.
    The key is never written into the code.
    """
    value = _clean_key_value(os.environ.get(CH_API_KEY_ENV, ""))
    if value:
        return value, "ortam degiskeni"

    for path in env_file_candidates():
        value = _clean_key_value(read_env_file(path).get(CH_API_KEY_ENV, ""))
        if value:
            if os.path.basename(path) != ".env":
                log.warning("Key read from '%s'. The file should be named '.env'; "
                            "Notepad on Windows appends '.txt' silently.",
                            os.path.basename(path))
            return value, path
    return "", ""


def describe_env_search():
    """
    Describe what the key search actually found, line by line, instead of
    just reporting that nothing was found.
    """
    lines = []
    raw = os.environ.get(CH_API_KEY_ENV)
    if raw is None:
        lines.append("  [none] {} is not set".format(CH_API_KEY_ENV))
    elif not _clean_key_value(raw):
        lines.append("  [EMPTY] {} is set but its value is empty".format(CH_API_KEY_ENV))
    else:
        lines.append("  [FOUND] {} environment variable".format(CH_API_KEY_ENV))

    seen_any_file = False
    for path in env_file_candidates():
        if not os.path.isfile(path):
            continue
        seen_any_file = True
        values = read_env_file(path)
        if CH_API_KEY_ENV in values and values[CH_API_KEY_ENV]:
            lines.append("  [FOUND] {}".format(path))
        elif values:
            lines.append("  [!]    {} exists but has no {}. "
                         "Keys it does contain: {}"
                         .format(path, CH_API_KEY_ENV, ", ".join(sorted(values)) or "(none)"))
        else:
            lines.append("  [!]    {} exists but has no readable KEY=VALUE line"
                         .format(path))
    if not seen_any_file:
        lines.append("  [none] No .env file in these directories:")
        for directory in env_search_dirs():
            lines.append("           {}".format(directory))

    lookalikes = find_env_lookalikes()
    if lookalikes:
        lines.append("")
        lines.append("  >>> Files with 'env' in the name were found here:")
        for path in lookalikes:
            lines.append("        {}".format(path))
        lines.append("  >>> Windows Explorer HIDES extensions: what looks like '.env'")
        lines.append("      may be '.env.txt' on disk. Turn on View > File name")
        lines.append("      extensions and check.")
    return lines


# Notepad's Save As silently appends .txt, and Explorer hides extensions, so
# a file the user sees as '.env' is '.env.txt' on disk. Accept the common
# wrong names and say which one was used, rather than leaving someone to
# spend an hour on a filename.
ENV_FILE_NAMES = [".env", ".env.txt", "env.txt", "env", ".env.env"]


def env_search_dirs():
    """Directories to search for a .env file, without duplicates."""
    dirs = [os.getcwd(), os.path.dirname(os.path.abspath(__file__))]
    unique = []
    for path in dirs:
        path = os.path.abspath(path)
        if path not in unique:
            unique.append(path)
    return unique


def env_file_candidates():
    """Every path to try: each directory crossed with each candidate name."""
    paths = []
    for directory in env_search_dirs():
        for name in ENV_FILE_NAMES:
            candidate = os.path.join(directory, name)
            if candidate not in paths:
                paths.append(candidate)
    return paths


def find_env_lookalikes():
    """
    Files in the searched directories whose name contains 'env' but is not
    exactly '.env'. This is what catches the hidden '.txt' Windows adds.
    """
    hits = []
    for directory in env_search_dirs():
        try:
            entries = os.listdir(directory)
        except OSError:
            continue
        for entry in entries:
            if entry == ".env":
                continue
            if "env" in entry.lower() and len(entry) <= 24:
                full = os.path.join(directory, entry)
                if os.path.isfile(full):
                    hits.append(full)
    return sorted(set(hits))


def _lookalike_env_vars():
    """
    Catch spelling and case mistakes such as 'ch_api_key' or 'CH_APIKEY'
    and report them.
    """
    target = CH_API_KEY_ENV.replace("_", "").upper()
    hits = []
    for name in os.environ:
        flat = name.strip().replace("_", "").replace("-", "").upper()
        if name != CH_API_KEY_ENV and (flat == target or "COMPANIESHOUSE" in flat):
            hits.append(name)
    return sorted(hits)


def api_key_help():
    """Help text shown when no key was found, including the diagnostics."""
    similar = _lookalike_env_vars()
    lines = ["No Companies House API key found.", "", "SEARCH RESULT:"]
    lines.extend(describe_env_search())
    lines += [
        "",
        "HOW TO SET IT:",
        "  Windows, persistent : setx {} \"your_key\"".format(CH_API_KEY_ENV),
        "                        >>> THEN FULLY CLOSE AND REOPEN THE TERMINAL <<<",
        "                        setx only affects NEWLY started processes. VS Code",
        "                        inherits its environment at launch, so a new tab",
        "                        is not enough; restart VS Code itself.",
        "  Windows, session    : set {}=your_key      (do NOT add quotes)".format(CH_API_KEY_ENV),
        "  PowerShell          : $env:{}=\"your_key\"".format(CH_API_KEY_ENV),
        "  macOS / Linux       : export {}=\"your_key\"".format(CH_API_KEY_ENV),
        "",
        "Or create a file named '.env' in this folder containing one line:",
        "  {}=your_key".format(CH_API_KEY_ENV),
        "  (.env is in .gitignore, so it never reaches the repository)",
        "",
        "Get a key at https://developer.company-information.service.gov.uk/",
        "",
        "To verify: python email_diagnostics.py check",
    ]
    if similar:
        lines.insert(3, "  >>> Similarly named variable(s) FOUND: {}".format(", ".join(similar)))
        lines.insert(4, "  >>> The name must be exactly '{}'.".format(CH_API_KEY_ENV))
    return "\n".join(lines)


def mask_key(value):
    """Mask the key for logging: keep the first 4 and last 2 characters."""
    if not value:
        return "(none)"
    if len(value) <= 8:
        return value[:2] + "*" * (len(value) - 2)
    return "{}{}{}".format(value[:4], "*" * (len(value) - 6), value[-2:])


def normalize_regnum(raw):
    """
    A UK company number is 8 characters. Excel will turn '01234567' into
    the number 1234567, and the API then returns 404, so the leading zeros
    are restored. Lettered prefixes such as SC, NI, OC and FC are already
    8 characters and are left alone.
    """
    text = re.sub(r"[^A-Za-z0-9]", "", (raw or "")).upper()
    if not text:
        return ""
    if text.isdigit():
        return text.zfill(8)
    return text


# ====================================================================
# SECTION 8 - OFFICER MATCHING
# ====================================================================

def _officer_name_parts(officer):
    """
    Extract (forename, other_forenames, surname, display name) from an
    officer record. The structured name_elements field is preferred; failing
    that the "SURNAME, Forename" form is parsed. Returns None when no name
    can be derived, which is the case for corporate officers.
    """
    if not isinstance(officer, dict):
        return None

    elements = officer.get("name_elements")
    display = (officer.get("name") or "").strip()

    if isinstance(elements, dict) and elements.get("surname"):
        forename = normalize(elements.get("forename") or "")
        others = normalize(elements.get("other_forenames") or "")
        surname = normalize(elements.get("surname") or "")
        pretty_parts = [p for p in [(elements.get("forename") or "").strip(),
                                    (elements.get("other_forenames") or "").strip(),
                                    (elements.get("surname") or "").strip()] if p]
        pretty = " ".join(pretty_parts) or display
        return forename, others, surname, pretty

    if display and "," in display:
        surname_raw, forenames_raw = display.split(",", 1)
        surname = normalize(surname_raw)
        forenames = normalize(forenames_raw).split()
        if surname:
            forename = forenames[0] if forenames else ""
            others = " ".join(forenames[1:])
            pretty = " ".join([p for p in forenames_raw.split()] + [surname_raw.strip()])
            return forename, others, surname, pretty.strip() or display

    return None   # corporate officer, or a record we cannot parse


def _surname_match_level(contact_variants, officer_surnames):
    """2 = exact match, 1 = within one edit, 0 = no match."""
    best = 0
    for officer_surname in officer_surnames:
        flat_officer = letters_only(officer_surname.replace(" ", ""))
        if not flat_officer:
            continue
        for variant in contact_variants:
            if not variant:
                continue
            if variant == flat_officer:
                return 2
            if (len(variant) >= 4 and len(flat_officer) >= 4
                    and edit_distance(variant, flat_officer,
                                    max_distance=SURNAME_MAX_DISTANCE) <= SURNAME_MAX_DISTANCE):
                best = max(best, 1)
    return best


def _forename_match_level(contact_firsts, contact_initials, officer_forenames):
    """2 = name or nickname matched, 1 = initial only, 0 = no match."""
    officer_set = set()
    for value in officer_forenames:
        for token in normalize(value).split():
            token = letters_only(token)
            if token:
                officer_set.add(token)
    if not officer_set or not contact_firsts:
        return 0

    officer_expanded = set()
    for token in officer_set:
        officer_expanded.update(expand_nicknames(token))
    officer_expanded.update(officer_set)

    for first in contact_firsts:
        if first in officer_expanded:
            return 2
        for token in officer_expanded:
            if len(first) >= 4 and len(token) >= 4 and edit_distance(first, token, max_distance=1) <= 1:
                return 2

    initials = set(i[0] for i in contact_firsts if i) | set(contact_initials)
    for token in officer_expanded:
        if token and token[0] in initials:
            return 1
    return 0


def list_active_officers(officers, limit=ACTIVE_SUGGESTION_LIMIT):
    """
    Names of officers currently in post, for the suggestion column.

    Used when the contact could not be matched: knowing who is actually at
    the company turns an unresolved row into an actionable one.
    Corporate officers are included by their registered name.
    """
    names = []
    for officer in officers or []:
        if officer.get("resigned_on"):
            continue
        parsed = _officer_name_parts(officer)
        if parsed:
            pretty = parsed[3]
        else:
            pretty = (officer.get("name") or "").strip()
        role = (officer.get("officer_role") or "").replace("-", " ")
        if pretty:
            names.append("{} ({})".format(pretty, role) if role else pretty)
    if not names:
        return ""
    if len(names) > limit:
        extra = len(names) - limit
        names = names[:limit] + ["+{} more".format(extra)]
    return " | ".join(names)


def match_contact_to_officers(name_data, officers):
    """
    Match the contact against the officer list. The SURNAME IS THE ANCHOR:
    it must match first, and only then is the forename considered.

    Returns dict(status, officer_name, reason, confidence), where status is
    "active", "resigned", "possible_active", "possible_resigned" or "none".
    """
    empty = {"status": "none", "officer_name": "", "reason": None, "confidence": "none"}
    if not officers:
        return empty

    contact_surnames = set(name_data["surname_variants"])
    if name_data["surname"]:
        contact_surnames.add(name_data["surname"])
    contact_surnames = set(s for s in contact_surnames if s)
    if not contact_surnames:
        return empty

    contact_firsts = set(f for f in name_data["first_candidates"] if f)
    contact_initials = set(name_data["initials"])

    matches = []
    for officer in officers:
        parsed = _officer_name_parts(officer)
        if not parsed:
            continue
        forename, others, surname, pretty = parsed

        # former_names covers surnames changed by marriage and the like
        former_surnames, former_forenames = [], []
        for former in (officer.get("former_names") or []):
            if isinstance(former, dict):
                if former.get("surname"):
                    former_surnames.append(normalize(former["surname"]))
                if former.get("forenames"):
                    former_forenames.append(normalize(former["forenames"]))

        surname_level = _surname_match_level(contact_surnames, [surname] + former_surnames)
        if surname_level == 0:
            continue

        forename_level = _forename_match_level(
            contact_firsts, contact_initials, [forename, others] + former_forenames
        )

        if forename_level == 2:
            confidence = "confident"
        elif forename_level == 1:
            confidence = "possible"
        else:
            confidence = "possible"

        matches.append({
            "pretty": pretty,
            "key": letters_only(pretty.replace(" ", "")),
            "confidence": confidence,
            "surname_level": surname_level,
            "forename_level": forename_level,
            "resigned": bool(officer.get("resigned_on")),
        })

    if not matches:
        return empty

    rank = {"confident": 2, "possible": 1}
    best_rank = max(rank[m["confidence"]] for m in matches)
    finalists = [m for m in matches if rank[m["confidence"]] == best_rank]

    distinct_people = set(m["key"] for m in finalists)
    ambiguous = len(distinct_people) > 1

    # If the same person has both a resigned and an active record, ACTIVE wins
    chosen_key = sorted(distinct_people)[0]
    same_person = [m for m in finalists if m["key"] == chosen_key]
    is_active = any(not m["resigned"] for m in same_person)
    display_name = same_person[0]["pretty"]

    confident = (best_rank == 2) and not ambiguous
    if confident:
        status = "active" if is_active else "resigned"
    else:
        status = "possible_active" if is_active else "possible_resigned"

    reason = None
    if ambiguous:
        reason = RSN.MULTIPLE_OFFICERS
    elif best_rank == 1 and finalists[0]["forename_level"] == 0:
        reason = RSN.SURNAME_ONLY
    elif finalists[0]["forename_level"] == 2 and len(display_name.split()) >= 3:
        reason = RSN.WITH_MIDDLE_NAME

    return {"status": status, "officer_name": display_name,
            "reason": reason, "confidence": "confident" if confident else "possible"}


# ====================================================================
# SECTION 9 - ROW PROCESSING AND ORCHESTRATION
# ====================================================================

def build_row_context(row, index_map):
    """Build the analysis context for one row, leaving the original values intact."""
    raw_first = row[index_map["first_name"]]
    raw_last = row[index_map["last_name"]]
    raw_email = row[index_map["email"]]
    raw_company = row[index_map["company"]]
    raw_regnum = row[index_map["regnum"]]

    name_data = resolve_person_name(raw_first, raw_last)
    company_tokens = clean_company_name(raw_company)
    email_info = parse_email(raw_email)
    regnum = normalize_regnum(raw_regnum)

    return {
        "row": row,
        "name": name_data,
        "company_tokens": company_tokens,
        "company_candidates": generate_company_domain_candidates(company_tokens),
        "email": email_info,
        "regnum": regnum,
        "result": None,
        "reason": None,
        "officer_name": "",
        "officer_status": "not_checked",
        "company_name": "",
        "suggestions": "",
        "typo": None,
    }


def run_email_stage(context):
    """Run the email stage and report whether it settled the row."""
    email_info = context["email"]

    if email_info["status"] == "missing":
        context["result"] = R.MISSING_EMAIL
        context["reason"] = None
        return True
    if email_info["status"] == "malformed":
        context["result"] = R.MALFORMED_EMAIL
        context["reason"] = None
        return True

    outcome = detect_email_typo(email_info, context["name"], context["company_candidates"])
    context["typo"] = outcome
    context["reason"] = outcome["reason"]
    if outcome["terminal"]:
        context["result"] = outcome["result"]
        return True
    return False


def finalise_ch_first(context):
    """
    ch_first mode: the email check runs AFTER the Companies House result.

    The official name (forename, MIDDLE NAMES and surname) is merged into the
    matching pool only when the officer match is CONFIDENT. Judging an email
    against the wrong person's name would produce a confidently wrong verdict.

    Result priority:
        data problem > company-level failure > resigned > typo > active
    result_reason always comes from the email stage, so every row carries
    both the officer status and the email verdict.
    """
    ch_result = context["result"]
    status = context["officer_status"]

    # Only trust the official name when the match is confident
    if context["officer_name"] and status in ("active", "resigned"):
        verified = resolve_person_name(context["officer_name"], "")
        name = context["name"]
        name["first_candidates"] |= verified["first_candidates"]
        name["surname_variants"] |= verified["surname_variants"]
        if verified["surname"]:
            name["surname_variants"].add(verified["surname"])
        for middle in verified["middles"]:
            if middle and middle not in name["middles"]:
                name["middles"].append(middle)

    # The email stage runs for every row and sets result_reason.
    context["result"] = None
    terminal = run_email_stage(context)
    email_result = context["result"]

    if email_result in (R.MISSING_EMAIL, R.MALFORMED_EMAIL):
        context["result"] = email_result           # a data problem outranks everything
    elif ch_result in (R.COMPANY_DISSOLVED, R.COMPANY_NOT_FOUND, R.LOOKUP_FAILED,
                       R.MISSING_REGNUM, R.CH_SKIPPED):
        context["result"] = ch_result              # company-level answer is decisive
    elif status == "resigned":
        context["result"] = ch_result              # resignation already explains the bounce
    elif terminal:
        context["result"] = email_result           # active officer, bad address -> typo
    else:
        context["result"] = ch_result              # active / possible / no match


def fetch_company_data(client, company_numbers):
    """
    Fetch officer data (and optionally the company profile) for the distinct
    company numbers, in parallel. Each regnum is queried once.

    Returns {regnum: {"officers": [...], "profile": {...}, "error": None|str}}
    """
    results = {}
    if not company_numbers:
        return results

    total = len(company_numbers)
    log.info("Companies House: %s distinct companies, %s threads, ~%.1f requests/s",
             total, CH_WORKERS, CH_RATE_LIMIT_PER_SEC)
    estimated = total * (2 if FETCH_COMPANY_PROFILE else 1) / CH_RATE_LIMIT_PER_SEC
    log.info("Estimated time: ~%.1f minutes", estimated / 60.0)

    counter = {"done": 0}
    counter_lock = threading.Lock()
    fatal = {"error": None}

    def worker(company_number):
        if fatal["error"]:
            return company_number, {"officers": [], "profile": None, "error": "aborted"}
        entry = {"officers": [], "profile": None, "error": None}
        before = client.stats["requests"]
        try:
            if FETCH_COMPANY_PROFILE:
                entry["profile"] = client.get_company_profile(company_number)
            entry["officers"] = client.get_companies_house_officers(company_number)
        except CompaniesHouseAuthError as exc:
            fatal["error"] = str(exc)
            entry["error"] = "auth"
        except CompanyNotFound:
            entry["error"] = "not_found"
        except LookupFailed as exc:
            entry["error"] = "failed: {}".format(exc)
            log.warning("%s lookup failed: %s", company_number, exc)
        except Exception as exc:                      # unexpected response shape, etc.
            entry["error"] = "failed: {}".format(exc)
            log.warning("%s unexpected error: %s: %s",
                        company_number, type(exc).__name__, exc)

        spent = client.stats["requests"] - before
        if spent > 1:
            log.debug("  %s used %s requests (pagination or retries)",
                      company_number, spent)
        with counter_lock:
            counter["done"] += 1
            if counter["done"] % 25 == 0 or counter["done"] == total:
                log.info("  ... %s/%s companies queried", counter["done"], total)
        return company_number, entry

    with ThreadPoolExecutor(max_workers=CH_WORKERS) as pool:
        for company_number, entry in pool.map(worker, company_numbers):
            results[company_number] = entry

    # Quota report: the requests actually sent, not an estimate.
    stats = client.stats
    per_company = (float(stats["requests"]) / total) if total else 0.0
    log.info("Companies House: %s HTTP requests / %s companies  (%.2f per company)",
             stats["requests"], total, per_company)
    if stats["retries"] or stats["rate_limited"] or stats["failed"]:
        log.info("  of which %s retries, %s rate limited (429), %s failed",
                 stats["retries"], stats["rate_limited"], stats["failed"])
    served = stats.get("server_page_size") or 0
    if served:
        log.info("  Largest page the server returned: %s records (we asked for %s)",
                 served, CH_PAGE_SIZE)
    # Attribute the extra requests correctly: retries or pagination?
    extra = stats["requests"] - (total * (2 if FETCH_COMPANY_PROFILE else 1))
    if extra > 0:
        if stats["retries"] >= extra:
            log.warning("  %s extra requests came from RETRIES (timeout / 429 / 5xx).",
                        extra)
        else:
            log.warning("  %s extra requests: %s from retries, the rest from "
                        "PAGINATION - the officer list did not fit one page "
                        "(total_results counts resigned officers too).",
                        extra, stats["retries"])

    reasons = {}
    for company_number, entry in results.items():
        if entry.get("error"):
            key = entry["error"].split(":")[0] if entry["error"] == "not_found" else entry["error"]
            reasons[key] = reasons.get(key, 0) + 1
    if reasons:
        log.warning("  Failed lookups (cause -> count):")
        for reason, count in sorted(reasons.items(), key=lambda kv: -kv[1])[:8]:
            log.warning("    %-52s %s", reason[:52], count)

    if fatal["error"]:
        raise CompaniesHouseAuthError(fatal["error"])
    return results


def apply_companies_house(context, company_data):
    """Apply the Companies House result to the row."""
    regnum = context["regnum"]

    if not regnum:
        context["result"] = R.MISSING_REGNUM
        context["officer_status"] = "not_checked"
        return

    entry = company_data.get(regnum)
    if entry is None:
        context["result"] = R.LOOKUP_FAILED
        context["officer_status"] = "lookup_failed"
        context["reason"] = context["reason"] or RSN.API_ERROR
        return

    profile = entry.get("profile")
    if isinstance(profile, dict):
        # companyhouse_names holds the official registered name, followed by
        # any former names the company traded under. A company that rebranded
        # often keeps the old domain, so the previous names are worth seeing.
        names = [profile.get("company_name") or ""]
        for previous in (profile.get("previous_company_names") or []):
            if isinstance(previous, dict) and previous.get("name"):
                names.append(previous["name"])
        context["company_name"] = " | ".join(n for n in names if n)

    if entry["error"] == "not_found":
        context["result"] = R.COMPANY_NOT_FOUND
        context["officer_status"] = "not_found"
        return
    if entry["error"]:
        context["result"] = R.LOOKUP_FAILED
        context["officer_status"] = "lookup_failed"
        context["reason"] = RSN.API_ERROR
        return

    # A closed company is the strongest explanation for a bounce
    if isinstance(profile, dict):
        status = (profile.get("company_status") or "").lower()
        if status in ("dissolved", "liquidation", "receivership", "administration",
                      "converted-closed", "closed"):
            context["result"] = R.COMPANY_DISSOLVED
            context["officer_status"] = "not_checked"
            return

    officers = entry.get("officers") or []
    if not officers:
        context["result"] = R.NO_OFFICER
        context["officer_status"] = "not_found"
        return

    match = match_contact_to_officers(context["name"], officers)
    context["officer_name"] = match["officer_name"]

    # When nobody matched, the row is a dead end unless we say who IS at the
    # company. List the currently serving officers so the contact can be
    # replaced rather than just marked unresolved.
    if match["status"] == "none":
        context["suggestions"] = list_active_officers(officers)
    if match["reason"]:
        context["reason"] = match["reason"]

    mapping = {
        "active": (R.ACTIVE, "active"),
        "resigned": (R.RESIGNED, "resigned"),
        "possible_active": (R.POSSIBLE_ACTIVE, "possible_active"),
        "possible_resigned": (R.POSSIBLE_RESIGNED, "possible_resigned"),
    }
    if match["status"] in mapping:
        label, officer_status = mapping[match["status"]]
        context["result"] = "{}: {}".format(label, match["officer_name"])
        context["officer_status"] = officer_status
    else:
        context["result"] = R.NO_OFFICER
        context["officer_status"] = "not_found"


# ====================================================================
# SECTION 10 - WRITING THE OUTPUT
# ====================================================================

OUTPUT_COLUMNS = ["result", "result_reason", "ch_officer_name", "ch_officer_status",
                  "companyhouse_names", "active_officer_suggestions"]

DEBUG_COLUMNS = [
    "dbg_clean_first", "dbg_clean_middles", "dbg_clean_surname", "dbg_nicknames",
    "dbg_company_tokens", "dbg_email_local", "dbg_email_domain",
    "dbg_best_pattern", "dbg_best_distance", "dbg_domain_verdict", "dbg_regnum_used",
]


def _build_output_headers(original_headers):
    headers = list(original_headers) + list(OUTPUT_COLUMNS)
    if DEBUG:
        headers.extend(DEBUG_COLUMNS)
    return headers


def _build_output_row(context):
    values = list(context["row"])
    values.extend([
        context["result"] or "",
        context["reason"] or "",
        context["officer_name"] or "",
        context["officer_status"] or "",
        context["company_name"] or "",
        context["suggestions"] or "",
    ])
    if DEBUG:
        typo = context["typo"] or {}
        name = context["name"]
        values.extend([
            name["first"],
            " ".join(name["middles"]),
            name["surname"],
            " ".join(name["nicknames"]),
            " ".join(context["company_tokens"]),
            context["email"]["local"],
            context["email"]["domain"],
            typo.get("best_pattern", ""),
            typo.get("best_distance", ""),
            typo.get("domain_verdict", ""),
            context["regnum"],
        ])
    return values


def write_output(path, original_headers, contexts, meta=None):
    """
    Write the ONE output file, CSV or Excel according to its extension.
    The original columns are preserved exactly and the new ones appended.
    The input file is never modified.

    meta carries the delimiter and encoding of the input; CSV output reuses
    the same delimiter and is written as utf-8-sig so Excel opens it correctly.
    """
    headers = _build_output_headers(original_headers)
    meta = meta or {}

    try:
        if is_csv_path(path):
            sep = meta.get("delimiter") or ","
            # utf-8-sig: the BOM is what makes Excel read accented text correctly
            with io.open(path, "w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle, delimiter=sep, quoting=csv.QUOTE_MINIMAL)
                writer.writerow(headers)
                for context in contexts:
                    writer.writerow([("" if v is None else v)
                                     for v in _build_output_row(context)])
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "diagnostics"
            sheet.append(headers)
            for context in contexts:
                sheet.append(_build_output_row(context))
            sheet.freeze_panes = "A2"
            workbook.save(path)
    except IOError as exc:
        raise IOError(
            "Could not write the output file: {}\n"
            "It may be open in Excel; close it and try again.\nDetail: {}"
            .format(os.path.abspath(path), exc)
        )
    log.info("Output written: %s  (%s rows)", os.path.abspath(path), len(contexts))


def print_summary(contexts):
    """Print the result distribution to the console. No extra file is created."""
    counts = OrderedDict()
    for context in contexts:
        label = (context["result"] or "").split(":")[0].strip() or "(empty)"
        counts[label] = counts.get(label, 0) + 1

    log.info("-" * 52)
    log.info("RESULT DISTRIBUTION (%s rows)", len(contexts))
    for label, count in sorted(counts.items(), key=lambda kv: -kv[1]):
        share = 100.0 * count / len(contexts) if contexts else 0.0
        log.info("  %-38s %5s  (%4.1f%%)", label, count, share)
    log.info("-" * 52)


# ====================================================================
# SECTION 11 - MAIN
# ====================================================================

def main():
    start_time = time.time()
    log.info("email_diagnostics %s  (python %s)", __version__,
             ".".join(str(n) for n in sys.version_info[:3]))

    # --- Safety: never write over the input file ---
    if os.path.abspath(INPUT_FILE) == os.path.abspath(OUTPUT_FILE):
        raise ValueError("The output file cannot be the input file. The input must be preserved.")

    # --- Load and validate ---
    original_headers, normalized_headers, rows, meta = load_data(
        INPUT_FILE, INPUT_SHEET, INPUT_DELIMITER, INPUT_ENCODING)
    index_map = validate_columns(normalized_headers)

    # --- Filter by status ---
    problematic = filter_problematic_statuses(rows, index_map["status"])
    if MAX_ROWS:
        problematic = problematic[:MAX_ROWS]
        log.warning("--limit is set: only the first %s rows will be processed.", MAX_ROWS)
    if not problematic:
        log.warning("No rows have a bounce status. An empty output will be written.")
        write_output(OUTPUT_FILE, original_headers, [], meta)
        return

    # --- Clean and run the email stage ---
    contexts = [build_row_context(row, index_map) for row in problematic]

    pending = []
    for context in contexts:
        if LOOKUP_MODE == "ch_first":
            pending.append(context)                 # every row goes to the API
        else:
            if not run_email_stage(context):        # no typo, so consult the API
                pending.append(context)

    log.info("Email stage: %s of %s rows will go to Companies House.",
             len(pending), len(contexts))

    # --- Companies House ---
    # NOTE: --limit bounds ROWS, not companies. Ten rows carrying three
    # distinct regnums means only three companies are queried. Use
    # --limit-companies to bound the thing that costs quota.
    company_numbers = sorted(set(c["regnum"] for c in pending if c["regnum"]))

    if MAX_COMPANIES and len(company_numbers) > MAX_COMPANIES:
        kept = set(company_numbers[:MAX_COMPANIES])
        dropped = len(company_numbers) - len(kept)
        company_numbers = sorted(kept)
        skipped = 0
        for context in pending:
            if context["regnum"] not in kept:
                context["result"] = context["result"] or R.CH_SKIPPED
                context["officer_status"] = "not_checked"
                skipped += 1
        pending = [c for c in pending if c["regnum"] in kept]
        log.warning("--limit-companies %s: %s companies left out of scope "
                    "(%s rows marked '%s').",
                    MAX_COMPANIES, dropped, skipped, R.CH_SKIPPED)

    log.info("Scope: %s rows -> %s distinct companies "
             "(at least one request each)",
             len(pending), len(company_numbers))

    if DRY_RUN:
        # Companies House was skipped ON PURPOSE. That is not a failure, so
        # writing 'lookup_failed' would misrepresent it; a distinct value is
        # used instead. The reason from the email stage is kept, so the typo
        # analysis survives in full and only the officer check is missing.
        log.warning("Companies House skipped (--dry-run / --no-ch): %s rows "
                    "marked without an officer check.", len(pending))
        for context in pending:
            # In ch_first the email stage normally runs AFTER Companies House.
            # With the lookup skipped it can still run, which is the entire
            # point of --no-ch; otherwise every row would be a bare 'skipped'.
            if LOOKUP_MODE == "ch_first":
                run_email_stage(context)
            context["result"] = context["result"] or R.CH_SKIPPED
            context["officer_status"] = "not_checked"
    elif company_numbers or pending:
        api_key, key_source = get_api_key()
        if not api_key:
            raise EnvironmentError(api_key_help())
        log.info("Companies House key found (source: %s)", key_source)
        client = CompaniesHouseClient(api_key, CH_RATE_LIMIT_PER_SEC, CH_MAX_REQUESTS)
        company_data = fetch_company_data(client, company_numbers)

        for context in pending:
            apply_companies_house(context, company_data)
            if LOOKUP_MODE == "ch_first":
                finalise_ch_first(context)

    # Safety net: no row may be left without a result
    for context in contexts:
        if not context["result"]:
            context["result"] = R.NO_OFFICER

    # --verbose: log the decision for every row
    for context in contexts:
        log.debug("%-40s -> %-34s %s",
                  context["email"]["email"] or "(no email)",
                  context["result"], context["reason"] or "")

    # --- Write the output ---
    write_output(OUTPUT_FILE, original_headers, contexts, meta)
    print_summary(contexts)
    log.info("Finished in %.1f seconds", time.time() - start_time)


# ====================================================================
# SECTION 12 - SETUP CHECK  (the check command)
# ====================================================================

# The company number used to test the key. Which one does not matter:
# any response other than 401/403 - including a 404 - proves that
# authentication succeeded.
CHECK_TEST_COMPANY = "00000006"


def _mark(ok):
    return "[OK]  " if ok else "[FAIL]"


def check_setup(input_path=None, skip_api=False, delimiter=None, encoding=None):
    """
    Verify the setup end to end and say exactly where it breaks.

    In order: Python version, packages, TLS environment, API key, whether the
    key is actually accepted, and optionally whether the input file parses and
    has the required columns.

    Returns 0 when everything is ready, 1 otherwise.
    """
    problems = []

    print("=" * 68)
    print(" SETUP CHECK")
    print("=" * 68)

    # --- 1) Python ---
    version = ".".join(str(n) for n in sys.version_info[:3])
    ok_version = sys.version_info >= (3, 6)
    print("%s Python %s  (%s)" % (_mark(ok_version), version, sys.platform))
    if not ok_version:
        problems.append("Python 3.6 or newer is required.")

    # --- 2) Packages ---
    for module, name in [(openpyxl, "openpyxl"), (requests, "requests")]:
        installed = getattr(module, "__version__", "?")
        print("%s %-9s %s" % (_mark(True), name, installed))

    # --- 2b) TLS environment, which is what matters when a handshake fails ---
    try:
        import ssl as _ssl
        print("%s OpenSSL   %s" % (_mark(True), _ssl.OPENSSL_VERSION))
    except ImportError:
        pass
    for name in ("HTTPS_PROXY", "HTTP_PROXY", "REQUESTS_CA_BUNDLE", "SSL_CERT_FILE"):
        value = os.environ.get(name) or os.environ.get(name.lower())
        if value:
            print("       %-18s %s" % (name, value))
    if CH_CA_BUNDLE:
        print("       %-18s %s" % ("--ca-bundle", CH_CA_BUNDLE))
    if not CH_VERIFY_SSL:
        print("       %-18s %s" % ("--insecure", "TLS VERIFICATION IS OFF"))

    # --- 3) API key ---
    api_key, source = get_api_key()
    if api_key:
        print("%s Key found: %s   (source: %s)"
              % (_mark(True), mask_key(api_key), source))
    else:
        print("%s No key found" % _mark(False))
        for line in describe_env_search():
            print(line)
        problems.append("api_key")

    # --- 4) Is the key actually accepted? ---
    if api_key and not skip_api:
        print("      Sending one test request to Companies House...")
        client = CompaniesHouseClient(api_key)
        try:
            client.get_company_profile(CHECK_TEST_COMPANY)
            print("%s Key accepted, API access works." % _mark(True))
        except CompanyNotFound:
            # A 404 also proves that authentication succeeded
            print("%s Key accepted, API access works." % _mark(True))
        except CompaniesHouseAuthError:
            print("%s Key REJECTED (HTTP 401/403)." % _mark(False))
            print("      The key may have been copied incorrectly, or it may be a")
            print("      Test key where a Live key is needed. Check the Developer Hub.")
            problems.append("api_key_invalid")
        except SSLHandshakeError as exc:
            print("%s TLS handshake failed." % _mark(False))
            print()
            print(exc)
            problems.append("ssl")
        except LookupFailed as exc:
            print("%s Could not reach the API: %s" % (_mark(False), exc))
            print("      Could be connectivity, a proxy or a firewall.")
            print("      On a corporate network, set HTTPS_PROXY.")
            problems.append("api_unreachable")
    elif api_key and skip_api:
        print("      (--skip-api given, no test request sent)")

    # --- 5) Input file ---
    if input_path:
        print("-" * 68)
        print(" INPUT FILE: %s" % input_path)
        print("-" * 68)
        try:
            _orig, normalized, rows, _meta = load_data(input_path, None, delimiter, encoding)
            print("%s File read: %s rows" % (_mark(True), len(rows)))

            missing = [c for c in REQUIRED_COLUMNS if c not in normalized]
            if missing:
                print("%s Missing column(s): %s" % (_mark(False), ", ".join(missing)))
                print("      Columns in the file: %s"
                      % ", ".join(h for h in normalized if h))
                problems.append("columns")
            else:
                print("%s All required columns are present." % _mark(True))

                status_index = normalized.index("status")
                counts = {}
                for row in rows:
                    key = (row[status_index] or "(bos)").strip()
                    counts[key] = counts.get(key, 0) + 1
                problematic = sum(1 for row in rows
                                  if is_problematic_status(row[status_index]))
                print("%s Rows that will be analysed: %s / %s"
                      % (_mark(problematic > 0), problematic, len(rows)))
                if problematic == 0:
                    print("      No row is in the bounce family. Your status values:")
                    problems.append("no_rows")
                print("      Status distribution:")
                for value, count in sorted(counts.items(), key=lambda kv: -kv[1])[:12]:
                    flag = "analysed" if is_problematic_status(value) else "-"
                    print("        %-28s %5s  %s" % (value[:28], count, flag))

                regnum_index = normalized.index("regnum")
                empty_regnum = sum(1 for row in rows if not (row[regnum_index] or "").strip())
                if empty_regnum:
                    print("      Warning: regnum is empty on %s rows." % empty_regnum)
        except (IOError, ValueError) as exc:
            print("%s Could not read the file: %s" % (_mark(False), exc))
            problems.append("input")

    # --- Summary ---
    print("=" * 68)
    if not problems:
        print(" EVERYTHING IS READY. To run:")
        print("   python email_diagnostics.py triage --input list.csv"
              " --output result.csv --verbose")
        print("=" * 68)
        return 0

    print(" SOMETHING IS MISSING")
    print("=" * 68)
    if "api_key" in problems:
        print()
        print(api_key_help())
    return 1


# ====================================================================
# SECTION 13 - INSPECT A SINGLE COMPANY  (the inspect command)
# ====================================================================

# The fields the diagnosis ACTUALLY consumes. Everything else exists in the
# API but is not used here; the used ones are starred in the output.
USED_PROFILE_FIELDS = {"company_name", "company_status", "previous_company_names"}
USED_OFFICER_FIELDS = {"name", "name_elements", "former_names", "resigned_on", "officer_role"}


def _fmt_address(address):
    if not isinstance(address, dict):
        return ""
    parts = [address.get(k) for k in
             ("premises", "address_line_1", "address_line_2", "locality",
              "region", "postal_code", "country")]
    return ", ".join(p for p in parts if p)


def inspect_company(company_number, raw=False):
    """
    Show everything Companies House returns for one company number: the
    practical answer to "what data can I get from a regnum?".

    Costs two requests: the company profile and the officer list.
    """
    api_key, source = get_api_key()
    if not api_key:
        raise EnvironmentError(api_key_help())

    number = normalize_regnum(company_number)
    if number != str(company_number).strip():
        log.info("Company number normalised: %s -> %s", company_number, number)

    client = CompaniesHouseClient(api_key)
    profile = {}
    try:
        profile = client.get_company_profile(number)
    except CompanyNotFound:
        print("Company not found: {}".format(number))
        return 1
    officers = client.get_companies_house_officers(number)

    if raw:
        print(json.dumps({"profile": profile, "officers": officers},
                         indent=2, ensure_ascii=False, sort_keys=True))
        return 0

    def show(container, key, label, indent="  "):
        value = container.get(key)
        if value in (None, "", [], {}):
            return
        star = "*" if key in USED_PROFILE_FIELDS or key in USED_OFFICER_FIELDS else " "
        print("{}{} {:<26} {}".format(indent, star, key, value))

    print("=" * 72)
    print(" COMPANY  ({})".format(number))
    print("=" * 72)
    for key in ("company_name", "company_number", "company_status", "company_status_detail",
                "type", "jurisdiction", "date_of_creation", "date_of_cessation",
                "has_charges", "has_insolvency_history", "sic_codes"):
        show(profile, key, key)
    address = _fmt_address(profile.get("registered_office_address"))
    if address:
        print("    registered_office_address  {}".format(address))
    for previous in (profile.get("previous_company_names") or []):
        print("  * previous_company_names    {}  ({} -> {})".format(
            previous.get("name"), previous.get("effective_from"), previous.get("ceased_on")))

    print()
    print("=" * 72)
    print(" OFFICERS  ({} records)".format(len(officers)))
    print("=" * 72)
    for index, officer in enumerate(officers, start=1):
        parsed = _officer_name_parts(officer)
        status = "RESIGNED (%s)" % officer.get("resigned_on") if officer.get("resigned_on") else "ACTIVE"
        print("-" * 72)
        print(" #{}  {}   [{}]".format(index, officer.get("name") or "(no name)", status))
        print("-" * 72)
        elements = officer.get("name_elements")
        if isinstance(elements, dict):
            print("  * name_elements              title={} forename={} other={} surname={}".format(
                elements.get("title") or "-", elements.get("forename") or "-",
                elements.get("other_forenames") or "-", elements.get("surname") or "-"))
        elif parsed is None:
            print("    (corporate officer - no name_elements, skipped when matching)")
        for former in (officer.get("former_names") or []):
            print("  * former_names               {} {}".format(
                former.get("forenames") or "", former.get("surname") or ""))
        for key in ("officer_role", "appointed_on", "resigned_on", "nationality",
                    "occupation", "country_of_residence", "person_number"):
            show(officer, key, key)
        birth = officer.get("date_of_birth")
        if isinstance(birth, dict):
            print("    date_of_birth              {}/{}".format(
                birth.get("month"), birth.get("year")))
        address = _fmt_address(officer.get("address"))
        if address:
            print("    address                    {}".format(address))

    print()
    print("=" * 72)
    print(" Fields marked (*) are the ones this script uses for the diagnosis.")
    print(" For the untouched response: --raw")
    print(" Requests spent: {}".format(client.stats["requests"]))
    print("=" * 72)
    return 0


# ====================================================================
# SECTION 14 - COMMAND LINE INTERFACE
# ====================================================================

def build_arg_parser():
    """
    Usage:
        python email_diagnostics.py triage --input list.csv --output result.csv --verbose

    Input and output may each be .xlsx or .csv, chosen by extension.
    """
    parser = argparse.ArgumentParser(
        prog="email_diagnostics.py",
        description="Diagnose bounced email rows and write ONE output file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python email_diagnostics.py check\n"
            "  python email_diagnostics.py check --input list.csv\n"
            "  python email_diagnostics.py inspect 17107304\n"
            "  python email_diagnostics.py triage --input list.csv --output result.csv --verbose\n"
            "  python email_diagnostics.py triage -i list.xlsx -o result.xlsx --limit 50 --no-ch\n\n"
            "On a new machine run 'check' FIRST: it says what is missing.\n\n"
            "The API key is read from the {env} environment variable or a .env file.\n"
            "  Windows    : setx {env} \"your_key\"   then FULLY restart VS Code\n"
            "  PowerShell : $env:{env}=\"your_key\"\n"
            "  macOS/Linux: export {env}=\"your_key\"".format(env=CH_API_KEY_ENV)
        ),
    )
    parser.add_argument("--version", action="version",
                        version="email_diagnostics {}".format(__version__))
    subparsers = parser.add_subparsers(dest="command", metavar="command")

    inspect = subparsers.add_parser(
        "inspect",
        help="Show everything Companies House returns for one company number.",
        description="List every field the API returns for one company number. "
                    "Costs two requests.",
    )
    inspect.add_argument("regnum", help="Company number, e.g. 17107304")
    inspect.add_argument("--raw", action="store_true",
                         help="Print the untouched JSON response.")
    inspect.add_argument("--ca-bundle", default=_DEFAULTS["ca_bundle"], metavar="PATH",
                         help="Corporate root certificate (.pem).")
    inspect.add_argument("--insecure", action="store_true",
                         help="Turn OFF certificate verification. Unsafe; for diagnosis only.")
    inspect.add_argument("-v", "--verbose", action="store_true", help="Verbose logging.")

    check = subparsers.add_parser(
        "check",
        help="Verify the setup: packages, TLS, API key, input file.",
        description="Verify the setup end to end and say where it breaks. "
                    "Run this first on a new machine.",
    )
    check.add_argument("-i", "--input", default=None, metavar="PATH",
                       help="Also check the input file: does it parse, are the required "
                            "columns present, how many rows will be analysed.")
    check.add_argument("--skip-api", "--no-ch", "--dry-run", dest="skip_api",
                       action="store_true",
                       help="Do not send the test request; run fully offline.")
    check.add_argument("--delimiter", default=None, metavar="CHAR",
                       help="CSV delimiter (default: detected).")
    check.add_argument("--encoding", default=None, metavar="ENCODING",
                       help="CSV encoding (default: detected).")
    check.add_argument("--ca-bundle", default=_DEFAULTS["ca_bundle"], metavar="PATH",
                       help="Corporate root certificate (.pem).")
    check.add_argument("--insecure", action="store_true",
                       help="Turn OFF certificate verification. Unsafe; for diagnosis only.")
    check.add_argument("-v", "--verbose", action="store_true",
                       help="Verbose logging.")

    triage = subparsers.add_parser(
        "triage",
        help="Analyse the input file and write the result file.",
        description="Analyse the input file and write the result file.",
    )
    triage.add_argument("-i", "--input", default=_DEFAULTS["input"], metavar="PATH",
                        help="Input file (.xlsx / .xlsm / .csv / .tsv). "
                             "Default: %(default)s")
    triage.add_argument("-o", "--output", default=_DEFAULTS["output"], metavar="PATH",
                        help="Output file; CSV or Excel by extension. "
                             "Default: %(default)s")
    triage.add_argument("--sheet", default=_DEFAULTS["sheet"], metavar="NAME",
                        help="Excel sheet name (default: the first sheet).")
    triage.add_argument("--delimiter", default=None, metavar="CHAR",
                        help="CSV delimiter. Guessed from the header row if omitted "
                             "(one of ; , tab |).")
    triage.add_argument("--encoding", default=None, metavar="ENCODING",
                        help="CSV encoding. Tries utf-8-sig, cp1254, cp1252 in turn if omitted.")

    triage.add_argument("-v", "--verbose", action="store_true",
                        help="Log the decision for every row.")
    triage.add_argument("--debug", action="store_true",
                        help="Add the audit columns to the output file.")
    triage.add_argument("--dry-run", "--no-ch", "--skip-api", dest="dry_run",
                        action="store_true",
                        help="Never call Companies House; run the email analysis only. "
                             "Skipped rows are marked '{}'.".format(R.CH_SKIPPED))
    triage.add_argument("--limit", type=int, default=_DEFAULTS["limit"], metavar="N",
                        help="Process only the first N bounced ROWS.")

    triage.add_argument("--mode", choices=["typo_first", "ch_first"], default=_DEFAULTS["mode"],
                        help="ch_first: verify the name first, then judge the email. "
                             "typo_first: check for typos first and skip the API for those "
                             "rows. Default: %(default)s")
    triage.add_argument("--no-company-profile", dest="company_profile",
                        action="store_false", default=_DEFAULTS["company_profile"],
                        help="Skip the company profile request. Saves one request per "
                             "company but leaves companyhouse_names empty and cannot "
                             "detect dissolved companies.")
    triage.add_argument("--workers", type=int, default=_DEFAULTS["workers"], metavar="N",
                        help="Parallel Companies House threads. Default: %(default)s")
    triage.add_argument("--limit-companies", type=int, default=_DEFAULTS["limit_companies"], metavar="N",
                        help="Query at most N DISTINCT regnums. --limit bounds rows; this "
                             "bounds companies, which is what actually costs quota.")
    triage.add_argument("--max-requests", type=int, default=_DEFAULTS["max_requests"], metavar="N",
                        help="Hard ceiling on total HTTP requests. Once reached, no further "
                             "request is sent.")
    triage.add_argument("--ca-bundle", default=_DEFAULTS["ca_bundle"], metavar="PATH",
                        help="Corporate root certificate (.pem), for networks that intercept TLS.")
    triage.add_argument("--insecure", action="store_true",
                        help="Turn OFF certificate verification. Unsafe; for diagnosis only.")
    triage.add_argument("--rate", type=float, default=_DEFAULTS["rate"], metavar="N",
                        help="Requests per second. The documented limit is 600/5min = 2.0. "
                             "Default: %(default)s")
    return parser


def apply_cli_args(args):
    """Apply the command line arguments to the module settings."""
    global INPUT_FILE, OUTPUT_FILE, INPUT_SHEET, INPUT_DELIMITER, INPUT_ENCODING
    global DEBUG, DRY_RUN, MAX_ROWS, MAX_COMPANIES, LOOKUP_MODE, FETCH_COMPANY_PROFILE
    global CH_WORKERS, CH_RATE_LIMIT_PER_SEC, CH_MAX_REQUESTS
    global CH_CA_BUNDLE, CH_VERIFY_SSL

    INPUT_FILE = args.input
    OUTPUT_FILE = args.output
    INPUT_SHEET = args.sheet
    INPUT_DELIMITER = args.delimiter
    INPUT_ENCODING = args.encoding

    DEBUG = bool(args.debug)
    DRY_RUN = bool(args.dry_run)
    MAX_ROWS = args.limit
    MAX_COMPANIES = args.limit_companies
    LOOKUP_MODE = args.mode
    FETCH_COMPANY_PROFILE = bool(args.company_profile)
    CH_WORKERS = max(1, int(args.workers))
    CH_RATE_LIMIT_PER_SEC = max(0.1, float(args.rate))
    CH_MAX_REQUESTS = args.max_requests
    apply_tls_args(args)

    if args.verbose:
        log.setLevel(logging.DEBUG)
        logging.getLogger().setLevel(logging.DEBUG)
        log.debug("Settings: mode=%s workers=%s rate=%.1f/s debug=%s dry_run=%s limit=%s",
                  LOOKUP_MODE, CH_WORKERS, CH_RATE_LIMIT_PER_SEC, DEBUG, DRY_RUN, MAX_ROWS)


def apply_tls_args(args):
    """Apply --ca-bundle / --insecure; shared by every subcommand."""
    global CH_CA_BUNDLE, CH_VERIFY_SSL
    CH_CA_BUNDLE = getattr(args, "ca_bundle", None)
    CH_VERIFY_SSL = not getattr(args, "insecure", False)
    if CH_CA_BUNDLE and not os.path.isfile(CH_CA_BUNDLE):
        raise IOError("--ca-bundle file not found: {}".format(CH_CA_BUNDLE))


def cli(argv=None):
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    # No subcommand: show the help (subparsers are not required in Python 3.6)
    if not getattr(args, "command", None):
        parser.print_help()
        return 2

    if args.verbose:
        log.setLevel(logging.DEBUG)
        logging.getLogger().setLevel(logging.DEBUG)

    if args.command == "inspect":
        try:
            apply_tls_args(args)
            return inspect_company(args.regnum, args.raw)
        except (EnvironmentError, CompaniesHouseAuthError, LookupFailed,
                SSLHandshakeError, IOError) as error:
            log.error("%s", error)
            return 1
        except KeyboardInterrupt:
            log.error("Interrupted by the user.")
            return 130

    if args.command == "check":
        try:
            apply_tls_args(args)
            return check_setup(args.input, args.skip_api, args.delimiter, args.encoding)
        except KeyboardInterrupt:
            log.error("Interrupted by the user.")
            return 130

    apply_cli_args(args)
    try:
        main()
    except (ValueError, IOError, EnvironmentError, CompaniesHouseAuthError,
            SSLHandshakeError) as error:
        log.error("%s", error)
        return 1
    except KeyboardInterrupt:
        log.error("Interrupted by the user.")
        return 130
    return 0


if __name__ == "__main__":
    sys.exit(cli())
